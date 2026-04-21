"""Construction de l'onglet Operation."""

from typing import Dict, List
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from excel_utils import HYPERLINK_FONT, LABEL_FONT, THIN_BORDER, VALUE_FONT, format_display_value

_OPERATION_FIXED_ROWS_AFTER = {
    "montant_collecte": [
        ("Montant d'une obligation", "1"),
        ("Ticket Minimum", "1 000"),
    ]
}


def build_operation_sheet(wb: Workbook, results: Dict, fields: List[Dict], ws=None) -> int:
    """Cree l'onglet Operation avec les champs simples, groupes par section."""
    ws = ws or wb.active
    ws.title = "Opération"

    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 50

    sections = [
        ("emprunt", "Société emprunteuse"),
        ("operation", "Société opération"),
    ]

    op_fields = [
        f
        for f in fields
        if f.get("excel_sheet", "") not in ("{person_name}", "{company_name}") and f.get("type") != "table"
    ]

    row = 1
    filled = 0
    current_section = None

    for field in op_fields:
        field_id = field["field_id"]

        section_key = None
        for key, _title in sections:
            if field_id.endswith(f"_{key}"):
                section_key = key
                break

        if section_key and section_key != current_section:
            if current_section is not None:
                row += 1

            current_section = section_key
            section_title = dict(sections).get(section_key, "")
            section_cell = ws.cell(row=row, column=2, value=section_title)
            section_cell.font = Font(bold=True, size=12, color="FFFFFF")
            section_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            section_cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            row += 1

        if section_key is None and current_section is not None and current_section != "__other":
            row += 1
            other_cell = ws.cell(row=row, column=2, value="Opération")
            other_cell.font = Font(bold=True, size=12, color="FFFFFF")
            other_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            other_cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            row += 1
            current_section = "__other"

        label = field.get("label", field_id)
        value = results.get(field_id)

        ws.cell(row=row, column=2, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

        val_cell = ws.cell(row=row, column=3, value=format_display_value(value))
        val_cell.font = VALUE_FONT
        val_cell.border = THIN_BORDER
        val_cell.alignment = Alignment(wrap_text=True)

        if field_id == "localisation_projet" and value:
            val_cell.hyperlink = f"https://earth.google.com/web/search/{quote(value)}"
            val_cell.font = HYPERLINK_FONT

        if value:
            filled += 1

        for fixed_label, fixed_value in _OPERATION_FIXED_ROWS_AFTER.get(field_id, []):
            row += 1
            ws.cell(row=row, column=2, value=fixed_label).font = LABEL_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER

            fixed_val_cell = ws.cell(row=row, column=3, value=fixed_value)
            fixed_val_cell.font = VALUE_FONT
            fixed_val_cell.border = THIN_BORDER
            fixed_val_cell.alignment = Alignment(wrap_text=True)

        row += 1

    return filled
