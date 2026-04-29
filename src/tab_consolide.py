from io import BytesIO
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook
from openpyxl import Workbook
from copy import copy


def _copy_sheet(source_ws, target_wb, sheet_title: str) -> None:
    ws = target_wb.create_sheet(title=sheet_title)
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
            if cell.hyperlink:
                new_cell.hyperlink = cell.hyperlink

    for col_letter, col_dim in source_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = col_dim.width

    for row_num, row_dim in source_ws.row_dimensions.items():
        ws.row_dimensions[row_num].height = row_dim.height

    if source_ws.freeze_panes:
        ws.freeze_panes = source_ws.freeze_panes


def _unique_title(base: str, existing: set) -> str:
    if base not in existing:
        return base
    i = 2
    while f"{base} ({i})" in existing:
        i += 1
    return f"{base} ({i})"


def build_consolidated_excel(audit_bytes: bytes, immo_bytes: bytes) -> bytes:
    wb_audit = load_workbook(BytesIO(audit_bytes))
    wb_immo = load_workbook(BytesIO(immo_bytes))

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    used_titles: set = set()

    for ws in wb_audit.worksheets:
        title = _unique_title(ws.title, used_titles)
        used_titles.add(title)
        _copy_sheet(ws, wb_out, title)

    for ws in wb_immo.worksheets:
        title = _unique_title(ws.title, used_titles)
        used_titles.add(title)
        _copy_sheet(ws, wb_out, title)

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output.getvalue()


def render_consolide_tab() -> None:
    st.header("Export Consolidé")
    st.caption("Fusionne le rapport Audit et le comparatif Immobilier en un seul fichier Excel.")

    # ---- Audit Excel ----
    audit_path_str = st.session_state.get("latest_excel_path")
    audit_project = st.session_state.get("latest_excel_project_name", "")
    audit_bytes: bytes | None = None

    if audit_path_str and Path(audit_path_str).exists():
        with open(audit_path_str, "rb") as f:
            audit_bytes = f.read()
        st.success(f"Rapport Audit disponible : **{audit_project}**")
    else:
        st.warning("Aucun rapport Audit disponible. Lancez d'abord un audit dans l'onglet **Audit**.")

    # ---- Immo Excel ----
    immo_result = st.session_state.get("immo_result")
    immo_bytes: bytes | None = None

    if immo_result:
        from tab_immo import build_immo_excel_export
        immo_bytes = build_immo_excel_export(immo_result)
        st.success("Comparatif Immobilier disponible.")
    else:
        st.warning("Aucun comparatif Immobilier disponible. Lancez une recherche dans l'onglet **Comparateur**.")

    st.divider()

    if audit_bytes and immo_bytes:
        if st.button("Générer le fichier consolidé", type="primary"):
            with st.spinner("Fusion en cours..."):
                consolidated = build_consolidated_excel(audit_bytes, immo_bytes)
            filename = f"consolide_{audit_project}.xlsx" if audit_project else "consolide.xlsx"
            st.download_button(
                label="Télécharger le fichier consolidé",
                data=consolidated,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    else:
        st.info("Les deux fichiers doivent être générés pour pouvoir les fusionner.")
