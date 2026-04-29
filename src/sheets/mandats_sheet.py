"""Construction de l'onglet Mandats."""

import re
import unicodedata
from typing import Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from core.excel_utils import (
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    HYPERLINK_FONT,
    RADIATED_ROW_FILL,
    THIN_BORDER,
    VALUE_FONT,
    apply_numeric_format,
    format_display_value,
    format_number_with_spaces,
)
from core.normalization import canonical_name


def _slugify_company_name(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value


def _build_pappers_company_url(company: dict) -> Optional[str]:
    nom_societe = (company.get("nom_societe") or company.get("societe") or "").strip()
    siren = str(company.get("siren") or "").strip()
    if not nom_societe or not siren:
        return None

    slug = _slugify_company_name(nom_societe)
    if not slug:
        return None

    return f"https://www.pappers.fr/entreprise/{slug}-{siren}"


def _build_infos_financieres(company: dict) -> Optional[str]:
    parts = []

    if company.get("capital") is not None:
        parts.append(f"Capital: {format_number_with_spaces(company.get('capital'))}")
    if company.get("chiffre_affaires") is not None:
        parts.append(f"CA: {format_number_with_spaces(company.get('chiffre_affaires'))}")
    if company.get("resultat_net") is not None:
        parts.append(f"RN: {format_number_with_spaces(company.get('resultat_net'))}")

    return " | ".join(parts) or None


def _build_statut(company: dict) -> Optional[str]:
    parts = []
    if company.get("statut"):
        parts.append(company.get("statut"))
    if company.get("statut_rcs"):
        parts.append(f"RCS: {company.get('statut_rcs')}")
    return " | ".join(parts) or None


def _is_radiated_company(company: dict) -> bool:
    statut = canonical_name(str(company.get("statut") or ""))
    statut_rcs = canonical_name(str(company.get("statut_rcs") or ""))
    return "radie" in statut or "radie" in statut_rcs


def _merge_commentaires(*values) -> Optional[str]:
    merged = []
    seen = set()

    for value in values:
        text = re.sub(r"\s+", " ", str(value or "").strip())
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        merged.append(text)

    return "\n".join(merged) or None


def build_mandats_sheet(wb: Workbook, pappers_mandats: Dict) -> int:
    """Cree l'onglet Mandats avec les societes Pappers."""
    if not pappers_mandats:
        return 0

    ws = wb.create_sheet("Mandats")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 30

    headers = ["Société", "Rôle", "Activité", "Création", "Infos financières", "SIREN", "Statut", "Nb représentants", "Détention", "Commentaires"]
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=2 + col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    filled = 0
    current_row = 3

    for folder_name, folder_payload in pappers_mandats.items():
        if not folder_payload:
            continue

        folder_cell = ws.cell(row=current_row, column=2, value=folder_name)
        folder_cell.font = Font(bold=True, size=12, color="2F5496")
        current_row += 1

        def _write_company(row: int, company: dict) -> int:
            nonlocal filled

            nom = (company.get("nom_societe") or company.get("societe") or "").strip()
            statut_rcs = (company.get("statut_rcs") or "").strip().lower()
            row_fill = RADIATED_ROW_FILL if _is_radiated_company(company) else None
            role_value = company.get("role")
            if statut_rcs and statut_rcs != "inscrit" and not role_value:
                role_value = "X"

            cell = ws.cell(row=row, column=2, value=nom)
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

            url = _build_pappers_company_url(company)
            cell.font = HYPERLINK_FONT if url else VALUE_FONT
            if url:
                cell.hyperlink = url

            commentaires_value = company.get("commentaires") or _merge_commentaires(
                company.get("publication_difficulte_contenu"),
            )
            values = [
                role_value,
                company.get("activite"),
                company.get("date_creation") or company.get("creation"),
                _build_infos_financieres(company),
                str(company.get("siren") or "").strip() or None,
                _build_statut(company),
                company.get("nb_dirigeants_total"),
                company.get("detention"),
                commentaires_value,
            ]
            for col_idx, value in enumerate(values):
                display_value = format_display_value(value) if not isinstance(value, (int, float)) else value
                company_cell = ws.cell(row=row, column=3 + col_idx, value=display_value)
                company_cell.font = VALUE_FONT
                company_cell.border = THIN_BORDER
                if col_idx == len(values) - 1:
                    company_cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_fill:
                    company_cell.fill = row_fill
                apply_numeric_format(company_cell, value)
                if value:
                    filled += 1

            if nom:
                filled += 1
            return row + 1

        if isinstance(folder_payload, list):
            for company in folder_payload:
                current_row = _write_company(current_row, company)
            total_cell = ws.cell(row=current_row, column=2, value=f"Total : {len(folder_payload)} société(s)")
            total_cell.font = Font(bold=True, size=10)
            current_row += 2
            continue

        if isinstance(folder_payload, dict):
            for person_name, companies in folder_payload.items():
                if not companies:
                    continue

                person_cell = ws.cell(row=current_row, column=2, value=person_name)
                person_cell.font = Font(bold=True, italic=True, size=10)
                current_row += 1

                for company in companies:
                    current_row = _write_company(current_row, company)

                total_cell = ws.cell(row=current_row, column=2, value=f"Total : {len(companies)} société(s)")
                total_cell.font = Font(bold=True, italic=True, size=9)
                current_row += 2

        current_row += 1

    return filled
