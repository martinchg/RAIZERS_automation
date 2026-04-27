"""Construction de l'onglet Patrimoine."""

import json
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from core.excel_utils import (
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    LABEL_FONT,
    NUMBER_FORMAT_INTEGER,
    SECTION_FONT,
    THIN_BORDER,
    VALUE_FONT,
    apply_numeric_format,
    format_display_value,
    normalize_key,
    to_number,
)
from core.normalization import is_archived_path

# ---------------------------------------------------------------------------
# Mapping clé canonique → variantes LLM acceptées
# Même logique que _ACTIF_POSTE_MAP dans financial_mapping.py
# ---------------------------------------------------------------------------

_PATRIMOINE_COL_MAP: Dict[str, List[str]] = {
    "type_bien": [
        "type_bien", "typebien", "type_de_bien", "nature", "naturebien",
        "nature_bien", "type_immobilier", "typedebien",
        # ancien champ avant refactoring
        "type_detention",
    ],
    "adresse": [
        "adresse", "adresse_destination", "adresse_et_destination",
        "localisation", "adresse_complete",
    ],
    "surface": [
        "surface", "surface_habitable", "superficie", "superficie_habitable",
        "surface_m2", "surface_en_m2",
    ],
    "type_de_detention": [
        "type_de_detention", "type_detention", "mode_de_detention",
        "mode_detention", "forme_juridique", "detention",
    ],
    "pct_detention": [
        "pct_detention", "pourcentage_detention", "taux_detention",
        "quote_part", "quotepartdetenue", "de_detention",
    ],
    "valeur_acquisition": [
        "valeur_acquisition", "valeur_d_acquisition", "prix_acquisition",
        "cout_acquisition", "prix_d_acquisition", "valeur_achat",
    ],
    "valeur_bien": [
        "valeur_bien", "estimation_actuelle", "valeur_actuelle",
        "valeur_venale", "valeur_marche", "valeur_estimee",
    ],
    "valeur_banque": [
        "valeur_banque", "capital_restant_du", "encours_credit",
        "encours_pret", "capital_du", "restant_du",
    ],
    "valeur_nette_detenue": ["valeur_nette_detenue"],
    "garanties_donnees": [
        "garanties_donnees", "garanties", "type_garantie",
        "type_de_garantie", "surete",
    ],
    "revenus_locatifs": [
        "revenus_locatifs", "revenus_fonciers", "loyers",
        "loyers_annuels", "revenus_annuels",
    ],
    "periode_revenus": [
        "periode_revenus", "periode", "frequence_revenus",
    ],
    # champs pour les autres tables (sociétés, mandats…)
    "societe": ["societe", "raison_sociale", "denomination"],
    "activite": ["activite"],
    "creation": ["creation"],
    "infos_financieres": ["infos_financieres", "chiffres_cles"],
    "role_detention": ["role_detention"],
    "commentaires": ["commentaires"],
}

# Index inverse : variante normalisée → clé canonique
_VARIANT_TO_CANONICAL: Dict[str, str] = {}
for _canon, _variants in _PATRIMOINE_COL_MAP.items():
    for _v in _variants:
        _VARIANT_TO_CANONICAL[normalize_key(_v)] = _canon


# ---------------------------------------------------------------------------
# Résolution d'une valeur dans un dict de ligne LLM
# ---------------------------------------------------------------------------

def _resolve_table_value(row_data: dict, key: str):
    """Résout la valeur pour la clé canonique `key` depuis un dict LLM."""
    candidates = _PATRIMOINE_COL_MAP.get(key, [key])
    normalized_row = {normalize_key(str(k)): v for k, v in row_data.items()}

    # 1. Correspondance directe (exact ou normalisée)
    for candidate in candidates:
        if candidate in row_data and row_data[candidate] is not None:
            return row_data[candidate]
        nk = normalize_key(candidate)
        if nk in normalized_row and normalized_row[nk] is not None:
            return normalized_row[nk]

    # 2. Correspondance via index inverse (toute clé du dict → canonique → comparaison)
    for raw_key, raw_val in row_data.items():
        if raw_val is None:
            continue
        canon = _VARIANT_TO_CANONICAL.get(normalize_key(str(raw_key)))
        if canon == key:
            return raw_val

    # 3. Champ calculé
    if key == "valeur_nette_detenue":
        vb = to_number(_resolve_table_value(row_data, "valeur_bien"))
        vk = to_number(_resolve_table_value(row_data, "valeur_banque"))
        if vb is not None and vk is not None:
            net = vb - vk
            return int(net) if float(net).is_integer() else net

    return None


# ---------------------------------------------------------------------------
# Résolution des personnes
# ---------------------------------------------------------------------------

def resolve_person_sheets(
    results: Dict, person_folder_map: Optional[Dict[str, str]] = None
) -> List[Tuple[str, str]]:
    persons: List[Tuple[str, str]] = []

    if person_folder_map:
        for suffix, folder_name in sorted(person_folder_map.items()):
            if is_archived_path(folder_name):
                continue
            persons.append((suffix, folder_name))
        if persons:
            return persons

    for field_id, value in sorted(results.items()):
        if field_id.startswith("patrimoine_personne_nom") and value:
            suffix = field_id.replace("patrimoine_personne_nom", "")
            persons.append((suffix, value.strip()))

    return persons


# ---------------------------------------------------------------------------
# Construction de l'onglet
# ---------------------------------------------------------------------------

def build_patrimoine_sheet(
    wb: Workbook,
    results: Dict,
    fields: List[Dict],
    logger,
    person_folder_map: Optional[Dict[str, str]] = None,
) -> int:
    """Crée l'onglet Patrimoine avec toutes les personnes."""
    persons = resolve_person_sheets(results, person_folder_map=person_folder_map)
    if not persons:
        return 0

    ws = wb.create_sheet("Patrimoine")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 8
    ws.column_dimensions["J"].width = 25

    filled = 0
    current_row = 1

    simple_fields = [f for f in fields if f.get("excel_sheet") == "{person_name}" and f.get("type") != "table"]
    table_fields = [f for f in fields if f.get("excel_sheet") == "{person_name}" and f.get("type") == "table"]

    for suffix, person_name in persons:
        section_cell = ws.cell(row=current_row, column=2, value=person_name)
        section_cell.font = Font(bold=True, size=13, color="2F5496")
        current_row += 1

        for field in simple_fields:
            field_id = field["field_id"]
            actual_fid = field_id + suffix
            actual_value = results.get(actual_fid)

            label = field.get("label", field_id)
            ws.cell(row=current_row, column=2, value=label).font = LABEL_FONT
            ws.cell(row=current_row, column=2).border = THIN_BORDER

            val_cell = ws.cell(row=current_row, column=3, value=format_display_value(actual_value))
            val_cell.font = VALUE_FONT
            val_cell.border = THIN_BORDER
            if actual_value:
                filled += 1
            current_row += 1

        for field in table_fields:
            field_id = field["field_id"]
            actual_fid = field_id + suffix
            actual_value = results.get(actual_fid)

            if not actual_value:
                continue

            try:
                rows = json.loads(actual_value) if isinstance(actual_value, str) else actual_value
            except (json.JSONDecodeError, TypeError):
                logger.warning("  %s: valeur table non parseable", actual_fid)
                continue

            if not rows:
                continue

            current_row += 1
            ws.cell(row=current_row, column=2, value=field.get("label", field_id)).font = SECTION_FONT
            current_row += 1

            col_map = field.get("column_mapping", {})
            is_patrimoine_table = "valeur_bien" in col_map and "valeur_banque" in col_map
            display_headers = list(col_map.keys())
            if is_patrimoine_table and "valeur_nette_detenue" not in display_headers:
                insert_idx = (
                    display_headers.index("valeur_banque") + 1
                    if "valeur_banque" in display_headers
                    else len(display_headers)
                )
                display_headers.insert(insert_idx, "valeur_nette_detenue")

            # Labels d'en-tête lisibles
            _HEADER_LABELS = {
                "type_detention": "Type de bien",
                "type_bien": "Type de bien",
                "adresse": "Adresse et destination",
                "surface": "Surface (m²)",
                "type_de_detention": "Type de détention",
                "pct_detention": "% détention",
                "valeur_acquisition": "Valeur d'acquisition",
                "valeur_bien": "Estimation actuelle",
                "valeur_banque": "Capital restant dû",
                "valeur_nette_detenue": "Valeur nette détenue",
                "garanties_donnees": "Garanties données",
                "revenus_locatifs": "Revenus locatifs",
                "periode_revenus": "M/A",
            }
            for col_idx, header in enumerate(display_headers):
                label_text = _HEADER_LABELS.get(header, header.replace("_", " ").title())
                cell = ws.cell(row=current_row, column=2 + col_idx, value=label_text)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THIN_BORDER
            current_row += 1

            numeric_keys = {"surface", "valeur_acquisition", "valeur_bien", "valeur_banque", "valeur_nette_detenue", "revenus_locatifs"}
            data_start_row = current_row

            for row_data in rows:
                for col_idx, key in enumerate(display_headers):
                    cell = ws.cell(row=current_row, column=2 + col_idx)
                    cell.border = THIN_BORDER

                    if key == "valeur_nette_detenue" and is_patrimoine_table:
                        col_vb = col_map.get("valeur_bien", "E")
                        col_vk = col_map.get("valeur_banque", "F")
                        col_pct = col_map.get("pct_detention", "C")
                        cell.value = f"=({col_vb}{current_row}-{col_vk}{current_row})*{col_pct}{current_row}"
                        cell.font = VALUE_FONT
                        cell.number_format = NUMBER_FORMAT_INTEGER
                    else:
                        cell_value = _resolve_table_value(row_data, key)
                        if isinstance(cell_value, (int, float)):
                            cell.value = cell_value
                        else:
                            cell.value = format_display_value(cell_value)
                        cell.font = VALUE_FONT
                        apply_numeric_format(cell, cell_value)
                        if cell_value is not None:
                            filled += 1
                current_row += 1

            data_end_row = current_row - 1
            total_font = Font(bold=True, size=10)
            ws.cell(row=current_row, column=2, value="TOTAL").font = total_font
            ws.cell(row=current_row, column=2).border = THIN_BORDER
            for col_idx, key in enumerate(display_headers):
                cell = ws.cell(row=current_row, column=2 + col_idx)
                cell.border = THIN_BORDER
                if key in numeric_keys :
                    col_letter = get_column_letter(2 + col_idx)
                    cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                    cell.font = total_font
                    cell.number_format = NUMBER_FORMAT_INTEGER
            current_row += 1

        current_row += 2

    return filled
