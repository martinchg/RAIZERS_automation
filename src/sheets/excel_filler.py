import argparse
import json
import logging
import re
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook

from sheets.bilan_sheet import _build_bilan_sheet, _build_compte_resultat_sheet
from sheets.lots_sheet import build_lots_sheet
from sheets.mandats_sheet import build_mandats_sheet
from core.normalization import canonical_name, extract_person_folder
from sheets.operation_sheet import build_operation_sheet
from sheets.patrimoine_sheet import build_patrimoine_sheet
from extraction.question_config import load_questions_config

logger = logging.getLogger(__name__)


def _normalize_folder_name(value: str) -> str:
    """Cle compacte (alphanumerique) basee sur la canonicalisation partagee."""
    return re.sub(r"[^a-z0-9]+", "", canonical_name(value or ""))


def _derive_person_folder_map_from_manifest(manifest_path: Path) -> Dict[str, str]:
    if not manifest_path.exists():
        return {}

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        return {}

    folder_map: Dict[str, str] = {}
    counter = 0
    for doc in manifest.get("files", []):
        folder_name = extract_person_folder(doc.get("source_path", ""))
        if not folder_name:
            continue
        key = _normalize_folder_name(folder_name)
        if key not in folder_map:
            folder_map[f"__{counter}"] = folder_name
            counter += 1

    return folder_map


def fill_excel(
    results: Dict,
    fields: List[Dict],
    output_dir: Path,
    person_folder_map: Optional[Dict[str, str]] = None,
    pappers_mandats: Optional[Dict[str, List[dict]]] = None,
    include_operation: bool = True,
    include_patrimoine: bool = True,
    include_bilan: bool = True,
    include_compte_resultat: bool = True,
    include_lots: bool = True,
) -> Path:
    """Cree un Excel from scratch avec les donnees extraites."""
    output_path = output_dir / "rapport.xlsx"
    wb = Workbook()
    default_ws = wb.active

    if include_operation:
        build_operation_sheet(wb, results, fields)
    else:
        wb.remove(default_ws)

    if include_patrimoine:
        build_patrimoine_sheet(
            wb,
            results,
            fields,
            logger=logger,
            person_folder_map=person_folder_map,
        )
    if include_bilan:
        _build_bilan_sheet(wb, results, fields)
    if include_compte_resultat:
        _build_compte_resultat_sheet(wb, results)
    if include_lots:
        build_lots_sheet(wb, results, fields, logger_=logger)
    build_mandats_sheet(wb, pappers_mandats or {})

    if not wb.sheetnames:
        placeholder_ws = wb.create_sheet("Rapport")
        placeholder_ws["A1"] = "Aucun onglet selectionne."

    wb.save(str(output_path))
    logger.info("Excel genere: %s", output_path)
    return output_path


def fill_excel_template(
    results: Dict,
    fields: List[Dict],
    template_path: str,
    output_dir: Path,
    person_folder_map: Optional[Dict[str, str]] = None,
    pappers_mandats: Optional[Dict[str, List[dict]]] = None,
):
    return fill_excel(
        results,
        fields,
        output_dir,
        person_folder_map=person_folder_map,
        pappers_mandats=pappers_mandats,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Cree un Excel depuis extraction_results.json")
    root = Path(__file__).parent.parent.parent.resolve()
    parser.add_argument(
        "--results",
        default=str(root / "output" / "extraction_results.json"),
        help="Chemin vers extraction_results.json",
    )
    parser.add_argument(
        "--questions",
        default=None,
        help="Chemin vers un fichier de questions unique (sinon charge les fichiers split du dossier config)",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Dossier de sortie (par defaut: dossier du fichier extraction_results.json)",
    )
    args = parser.parse_args()

    results_path = Path(args.results)
    output_dir = Path(args.output_dir) if args.output_dir else results_path.parent
    manifest_path = output_dir / "manifest.json"

    if not results_path.exists():
        project_hint = output_dir.name if output_dir.name else None
        hint = (
            f" Lance d'abord `python run.py extract --project {project_hint}` puis relance `fill`."
            if project_hint
            else ""
        )
        raise FileNotFoundError(f"extraction_results.json introuvable: {results_path}.{hint}")

    extraction_data = json.loads(results_path.read_text(encoding="utf-8"))
    questions_data = load_questions_config(root / "config", explicit_path=args.questions)

    person_folder_map = extraction_data.get("person_folders")
    if not person_folder_map:
        person_folder_map = _derive_person_folder_map_from_manifest(manifest_path)

    mandats_results_path = output_dir / "mandats_results.json"
    pappers_mandats = None
    if mandats_results_path.exists():
        mandats_data = json.loads(mandats_results_path.read_text(encoding="utf-8"))
        pappers_mandats = mandats_data.get("societes_par_personne")

    fields = [f for f in questions_data["fields"] if isinstance(f, dict) and f.get("field_id")]

    fill_excel(
        results=extraction_data["results"],
        fields=fields,
        output_dir=output_dir,
        person_folder_map=person_folder_map,
        pappers_mandats=pappers_mandats,
    )


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s — %(levelname)s — %(message)s",
    )
    main()
