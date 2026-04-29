"""Construction des onglets Bilan et Compte de resultat."""

from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font

from core.excel_utils import (
    BOLD_FONT,
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    ITALIC_FONT,
    NUMBER_FORMAT_INTEGER,
    RED_FONT,
    SECTION_FONT,
    THIN_BORDER,
    VALUE_FONT,
    apply_numeric_format,
    format_display_value,
    to_number,
)
from financial.financial_mapping import build_financial_table_lookup, resolve_financial_metric_value

_BILAN_ACTIF_ROWS = (
    ("Immobilisations corporelles", "immobilisations_corporelles", VALUE_FONT),
    ("Immobilisations financières", "immobilisations_financieres", VALUE_FONT),
    ("Stocks", "stocks", VALUE_FONT),
    ("Créances", "creances", VALUE_FONT),
    ("Trésorerie", "tresorerie", VALUE_FONT),
    ("Autres éléments d'actif", "autres_actif", VALUE_FONT),
)

_BILAN_PASSIF_DETAIL_ROWS = (
    ("          Capital social", "capital_social", ITALIC_FONT),
    ("          Résultat", "resultat_exercice", ITALIC_FONT),
)

_BILAN_PASSIF_ROWS = (
    ("Capitaux propres", "capitaux_propres", BOLD_FONT),
    ("Dettes financières", "dettes_financieres", VALUE_FONT),
    ("Dettes d'exploitation", "dettes_exploitation", VALUE_FONT),
    ("Dettes diverses", "dettes_diverses", VALUE_FONT),
    ("Autres éléments de passif", "autres_passif", VALUE_FONT),
)

_COMPTE_RESULTAT_ROWS = (
    ("Chiffre d'affaires", "chiffre_affaires", VALUE_FONT),
    ("Charges", "charges", VALUE_FONT),
    ("Salaires et charges sociales", "salaires_charges_sociales", VALUE_FONT),
    ("Impôts et taxes", "impots_taxes", VALUE_FONT),
    ("Dotations aux amortissements", "dotations", VALUE_FONT),
    ("Autres éléments", "autres_elements", VALUE_FONT),
    ("Résultat d'exploitation", "resultat_exploitation", BOLD_FONT),
    ("Résultat financier", "resultat_financier", VALUE_FONT),
    ("Résultat exceptionnel", "resultat_exceptionnel", VALUE_FONT),
    ("Impôts sur les sociétés", "impots_sur_les_societes", VALUE_FONT),
    ("Résultat net", "resultat_net", BOLD_FONT),
)


def _resolve_company_sheets(results: Dict) -> List[Tuple[str, str]]:
    companies: List[Tuple[str, str]] = []
    for field_id, value in sorted(results.items()):
        if field_id.startswith("bilan_societe_nom") and value:
            suffix = field_id.replace("bilan_societe_nom", "")
            companies.append((suffix, value.strip()))
    return companies


def _configure_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30


def _write_headers(ws, row: int, date_n, date_n1) -> int:
    col_n_label = f"Au {date_n}" if date_n else "Exercice N"
    col_n1_label = f"Au {date_n1}" if date_n1 else "Exercice N-1"
    for offset, header in enumerate(("En €", col_n_label, col_n1_label, "Commentaires"), start=2):
        cell = ws.cell(row=row, column=offset, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    return row + 1


def _write_metric_row(
    ws,
    row: int,
    label: str,
    key: Optional[str],
    font,
    results: Dict,
    suffix: str,
    table_lookup: Dict[str, Dict],
    statement_type: str,
) -> int:
    filled = 0
    label_cell = ws.cell(row=row, column=2, value=label)
    label_cell.font = font
    label_cell.border = THIN_BORDER

    for column, period in ((3, "n"), (4, "n1")):
        cell = ws.cell(row=row, column=column)
        cell.border = THIN_BORDER
        cell.font = font

        value = (
            resolve_financial_metric_value(results, suffix, table_lookup, key, period, statement_type)
            if key
            else None
        )
        number = to_number(value)
        if number is not None:
            cell.value = number
            apply_numeric_format(cell, number)
            filled += 1
        elif value:
            cell.value = format_display_value(value)
            filled += 1

    comment_cell = ws.cell(row=row, column=5)
    comment_cell.border = THIN_BORDER
    return filled


def _write_formula_row(
    ws,
    row: int,
    label: str,
    font,
    formula_n: str,
    formula_n1: str,
    percent: bool = False,
) -> None:
    label_cell = ws.cell(row=row, column=2, value=label)
    label_cell.font = font
    label_cell.border = THIN_BORDER

    for column, formula in ((3, formula_n), (4, formula_n1)):
        cell = ws.cell(row=row, column=column, value=formula)
        cell.font = font
        cell.border = THIN_BORDER
        cell.number_format = "0%" if percent else NUMBER_FORMAT_INTEGER

    ws.cell(row=row, column=5).border = THIN_BORDER


def _combined_bilan_lookup(results: Dict, suffix: str) -> Dict[str, Dict]:
    lookup = build_financial_table_lookup(results.get(f"bilan_actif_table{suffix}"), "bilan_actif_table")
    lookup.update(
        build_financial_table_lookup(results.get(f"bilan_passif_table{suffix}"), "bilan_passif_table")
    )
    return lookup


def _build_bilan_sheet(wb: Workbook, results: Dict, fields: List[Dict]):
    companies = _resolve_company_sheets(results)
    if not companies:
        return 0

    ws = wb.create_sheet("Bilan")
    _configure_sheet(ws)

    filled = 0
    row = 1

    for company_index, (suffix, company_name) in enumerate(companies):
        if company_index > 0:
            row += 3

        title_cell = ws.cell(row=row, column=2, value=company_name)
        title_cell.font = Font(bold=True, size=13, color="2F5496")
        row += 1

        row = _write_headers(
            ws,
            row,
            results.get(f"bilan_date_arrete_n{suffix}", ""),
            results.get(f"bilan_date_arrete_n1{suffix}", ""),
        )

        table_lookup = _combined_bilan_lookup(results, suffix)
        actif_start_row = row
        for label, key, font in _BILAN_ACTIF_ROWS:
            filled += _write_metric_row(ws, row, label, key, font, results, suffix, table_lookup, "bilan")
            row += 1
        actif_end_row = row - 1

        _write_formula_row(
            ws,
            row,
            "Total Actif",
            BOLD_FONT,
            f"=SUM(C{actif_start_row}:C{actif_end_row})",
            f"=SUM(D{actif_start_row}:D{actif_end_row})",
        )
        total_actif_row = row
        row += 1

        filled += _write_metric_row(
            ws,
            row,
            "Total Actif (source doc)",
            "total_actif",
            ITALIC_FONT,
            results,
            suffix,
            table_lookup,
            "bilan",
        )
        total_actif_source_row = row
        row += 1

        for label, key, font in _BILAN_PASSIF_DETAIL_ROWS:
            filled += _write_metric_row(ws, row, label, key, font, results, suffix, table_lookup, "bilan")
            row += 1

        passif_rows = {}
        for label, key, font in _BILAN_PASSIF_ROWS:
            passif_rows[key] = row
            filled += _write_metric_row(ws, row, label, key, font, results, suffix, table_lookup, "bilan")
            row += 1

        _write_formula_row(
            ws,
            row,
            "Total Passif",
            BOLD_FONT,
            (
                f"=C{passif_rows['capitaux_propres']}+C{passif_rows['dettes_financieres']}"
                f"+C{passif_rows['dettes_exploitation']}+C{passif_rows['dettes_diverses']}"
                f"+C{passif_rows['autres_passif']}"
            ),
            (
                f"=D{passif_rows['capitaux_propres']}+D{passif_rows['dettes_financieres']}"
                f"+D{passif_rows['dettes_exploitation']}+D{passif_rows['dettes_diverses']}"
                f"+D{passif_rows['autres_passif']}"
            ),
        )
        total_passif_row = row
        row += 1

        filled += _write_metric_row(
            ws,
            row,
            "Total Passif (source doc)",
            "total_passif",
            ITALIC_FONT,
            results,
            suffix,
            table_lookup,
            "bilan",
        )
        total_passif_source_row = row
        row += 2

        check_labels = (
            ("Check équilibre", f"=C{total_actif_row}=C{total_passif_row}", f"=D{total_actif_row}=D{total_passif_row}"),
            (
                "Check total actif source",
                f'=IF(C{total_actif_source_row}="",TRUE,C{total_actif_row}=C{total_actif_source_row})',
                f'=IF(D{total_actif_source_row}="",TRUE,D{total_actif_row}=D{total_actif_source_row})',
            ),
            (
                "Check total passif source",
                f'=IF(C{total_passif_source_row}="",TRUE,C{total_passif_row}=C{total_passif_source_row})',
                f'=IF(D{total_passif_source_row}="",TRUE,D{total_passif_row}=D{total_passif_source_row})',
            ),
            (
                "Check résultat",
                (
                    f"=C{passif_rows['capitaux_propres']}="
                    f"(C{passif_rows['capitaux_propres'] - 2}+C{passif_rows['capitaux_propres'] - 1})"
                ),
                (
                    f"=D{passif_rows['capitaux_propres']}="
                    f"(D{passif_rows['capitaux_propres'] - 2}+D{passif_rows['capitaux_propres'] - 1})"
                ),
            ),
        )
        for label, formula_n, formula_n1 in check_labels:
            label_cell = ws.cell(row=row, column=2, value=label)
            label_cell.font = RED_FONT
            label_cell.border = THIN_BORDER
            for column, formula in ((3, formula_n), (4, formula_n1)):
                cell = ws.cell(row=row, column=column, value=formula)
                cell.font = RED_FONT
                cell.border = THIN_BORDER
            row += 1

        row += 1

    return filled


def _build_compte_resultat_sheet(wb: Workbook, results: Dict):
    companies = _resolve_company_sheets(results)
    if not companies:
        return 0

    ws = wb.create_sheet("Compte de résultat")
    _configure_sheet(ws)

    filled = 0
    row = 1

    for company_index, (suffix, company_name) in enumerate(companies):
        if company_index > 0:
            row += 3

        title_cell = ws.cell(row=row, column=2, value=company_name)
        title_cell.font = Font(bold=True, size=13, color="2F5496")
        row += 1

        row = _write_headers(
            ws,
            row,
            results.get(f"bilan_date_arrete_n{suffix}", ""),
            results.get(f"bilan_date_arrete_n1{suffix}", ""),
        )

        table_lookup = build_financial_table_lookup(
            results.get(f"bilan_compte_resultat_table{suffix}"),
            "bilan_compte_resultat_table",
        )

        row_positions = {}
        for label, key, font in _COMPTE_RESULTAT_ROWS:
            row_positions[key] = row
            filled += _write_metric_row(
                ws,
                row,
                label,
                key,
                font,
                results,
                suffix,
                table_lookup,
                "compte_resultat",
            )
            row += 1

            if key in {"resultat_exploitation", "resultat_net"}:
                chiffre_affaires_row = row_positions["chiffre_affaires"]
                formula_row = row
                _write_formula_row(
                    ws,
                    formula_row,
                    "En % du CA",
                    ITALIC_FONT,
                    f"=IFERROR(C{row_positions[key]}/C{chiffre_affaires_row},0)",
                    f"=IFERROR(D{row_positions[key]}/D{chiffre_affaires_row},0)",
                    percent=True,
                )
                row += 1

    return filled


# ── Onglets 4 colonnes (PDF N / PDF N-1 / Pappers N / Pappers N-1) ────────────

def _row_font_4col(label: str) -> Font:
    import unicodedata
    nfd = unicodedata.normalize("NFD", label.lower().strip())
    canon = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    if canon.startswith("total") or canon in ("benefice ou perte", "resultat net"):
        return BOLD_FONT
    return VALUE_FONT


def _fmt_date(d: Optional[str], fallback: str) -> str:
    if not d:
        return fallback
    from datetime import datetime
    try:
        return f"Au {datetime.strptime(d, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    except ValueError:
        return f"Au {d}"


def _write_4col_headers(ws, row: int, dates: Dict[str, Optional[str]]) -> int:
    headers = [
        "En €",
        _fmt_date(dates.get("pdf_n"),     "PDF N"),
        _fmt_date(dates.get("pdf_n1"),    "PDF N-1"),
        _fmt_date(dates.get("pappers_n"), "Pappers N"),
        _fmt_date(dates.get("pappers_n1"), "Pappers N-1"),
        "Commentaires",
    ]
    for offset, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=offset, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    return row + 1


def _write_4col_rows(ws, row: int, section_rows: List[Dict[str, Any]]) -> int:
    for data_row in section_rows:
        label = str(data_row.get("label") or "").strip()
        if not label:
            continue
        is_computed = data_row.get("_computed", False)
        is_percent = data_row.get("_percent", False)
        font = ITALIC_FONT if is_computed else _row_font_4col(label)
        label_cell = ws.cell(row=row, column=2, value=label)
        label_cell.font = font
        label_cell.border = THIN_BORDER

        for col, key in ((3, "pdf_n"), (4, "pdf_n1"), (5, "pappers_n"), (6, "pappers_n1")):
            cell = ws.cell(row=row, column=col)
            cell.font = font
            cell.border = THIN_BORDER
            value = data_row.get(key)
            number = to_number(value)
            if number is not None:
                cell.value = number
                if is_percent:
                    cell.number_format = "0.0%"
                else:
                    apply_numeric_format(cell, number)
            elif value is not None:
                cell.value = format_display_value(value)

        ws.cell(row=row, column=7).border = THIN_BORDER
        row += 1
    return row


def _enrich_cdr_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Injecte les lignes calculées EBITDA et % CA après les lignes clés du CDR."""
    def _get(label: str) -> Dict:
        for r in rows:
            if r.get("label") == label:
                return r
        return {}

    def _add(a, b):
        if a is None and b is None:
            return None
        return (a or 0) + (b or 0)

    def _pct(num, denom):
        if num is None or not denom:
            return None
        return num / denom

    ca = _get("Chiffre d'affaires")
    rex = _get("Résultat d'exploitation")
    dotations = _get("Dotations aux amortissements")

    enriched = []
    for row in rows:
        enriched.append(row)
        label = row.get("label", "")

        if label == "Résultat d'exploitation":
            enriched.append({
                "label": "EBITDA",
                "pdf_n":      _add(rex.get("pdf_n"),     dotations.get("pdf_n")),
                "pdf_n1":     _add(rex.get("pdf_n1"),    dotations.get("pdf_n1")),
                "pappers_n":  _add(rex.get("pappers_n"), dotations.get("pappers_n")),
                "pappers_n1": _add(rex.get("pappers_n1"),dotations.get("pappers_n1")),
                "_computed": True,
            })
            enriched.append({
                "label": "En % du CA",
                "pdf_n":      _pct(row.get("pdf_n"),     ca.get("pdf_n")),
                "pdf_n1":     _pct(row.get("pdf_n1"),    ca.get("pdf_n1")),
                "pappers_n":  _pct(row.get("pappers_n"), ca.get("pappers_n")),
                "pappers_n1": _pct(row.get("pappers_n1"),ca.get("pappers_n1")),
                "_computed": True,
                "_percent": True,
            })

        elif label == "Résultat net":
            enriched.append({
                "label": "En % du CA",
                "pdf_n":      _pct(row.get("pdf_n"),     ca.get("pdf_n")),
                "pdf_n1":     _pct(row.get("pdf_n1"),    ca.get("pdf_n1")),
                "pappers_n":  _pct(row.get("pappers_n"), ca.get("pappers_n")),
                "pappers_n1": _pct(row.get("pappers_n1"),ca.get("pappers_n1")),
                "_computed": True,
                "_percent": True,
            })

    return enriched


def _configure_4col_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 38
    for col in ("C", "D", "E", "F"):
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["G"].width = 28


def build_bilan_4col_sheets(wb: Workbook, bilan_results: List[Dict[str, Any]]) -> None:
    """Construit les onglets Bilan 4col et Compte de résultat 4col.

    Args:
        bilan_results: liste de résultats du financial_bilan_integrator.run_for_pdf()
    """
    if not bilan_results:
        return

    ws_bilan = wb.create_sheet("Bilan")
    ws_cr    = wb.create_sheet("Compte de résultat")
    for ws in (ws_bilan, ws_cr):
        _configure_4col_sheet(ws)

    bilan_row = 1
    cr_row = 1

    for i, result in enumerate(bilan_results):
        company = result.get("company") or f"Société {i + 1}"
        sections = result.get("sections") or {}
        dates = result.get("dates") or {}

        if i > 0:
            bilan_row += 2
            cr_row += 2

        # — Bilan —
        title_cell = ws_bilan.cell(row=bilan_row, column=2, value=company)
        title_cell.font = Font(bold=True, size=13, color="2F5496")
        bilan_row += 1

        bilan_row = _write_4col_headers(ws_bilan, bilan_row, dates)

        for section_name, section_title in (("actif", "Actif"), ("passif", "Passif")):
            sect_cell = ws_bilan.cell(row=bilan_row, column=2, value=section_title)
            sect_cell.font = SECTION_FONT
            bilan_row += 1
            bilan_row = _write_4col_rows(ws_bilan, bilan_row, sections.get(section_name) or [])
            bilan_row += 1

        # — Compte de résultat —
        cr_title_cell = ws_cr.cell(row=cr_row, column=2, value=company)
        cr_title_cell.font = Font(bold=True, size=13, color="2F5496")
        cr_row += 1

        cr_row = _write_4col_headers(ws_cr, cr_row, dates)
        cdr_rows = _enrich_cdr_rows(sections.get("compte_resultat") or [])
        cr_row = _write_4col_rows(ws_cr, cr_row, cdr_rows)
        cr_row += 1
