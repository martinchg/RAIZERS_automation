from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from core.excel_utils import (
    BOLD_FONT,
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    LABEL_FONT,
    RED_FONT,
    THIN_BORDER,
    VALUE_FONT,
    apply_numeric_format,
    to_number,
)


SUMMARY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

EXPORT_COLUMNS = [
    ("source", "Source"),
    ("localisation", "Localisation"),
    ("property_type", "Type de bien"),
    ("prix_bas_m2", "Prix bas au m²"),
    ("prix_moyen_m2", "Prix moyen au m²"),
    ("prix_haut_m2", "Prix haut au m²"),
    ("status", "Statut"),
    ("url_source", "URL source"),
    ("method_used", "Méthode"),
    ("error", "Erreur"),
]


def _normalize_status(row: Dict[str, Any]) -> str:
    explicit = row.get("status")
    if explicit:
        return str(explicit).strip().lower()
    if row.get("error"):
        return "error"
    if any(row.get(key) is not None for key in ("prix_bas_m2", "prix_moyen_m2", "prix_haut_m2")):
        return "ok"
    return "empty"


def _coalesce_price(*values: Any) -> Optional[float]:
    for value in values:
        number = to_number(value)
        if number is not None:
            return float(number)
    return None


def _pick_price_block(raw: Dict[str, Any], property_type: Optional[str]) -> Dict[str, Any]:
    blocks = raw.get("prix")
    if not isinstance(blocks, list) or not blocks:
        return raw

    if property_type:
        target = str(property_type).strip().lower()
        for block in blocks:
            block_type = str(block.get("type_bien") or "").strip().lower()
            if block_type == target:
                return block

    if len(blocks) == 1 and isinstance(blocks[0], dict):
        return blocks[0]

    for block in blocks:
        if isinstance(block, dict) and any(
            block.get(key) is not None
            for key in ("prix_m2_min", "prix_m2_moyen", "prix_m2_max")
        ):
            return block

    return raw


def normalize_scraping_result(
    raw: Dict[str, Any],
    *,
    property_type: Optional[str] = None,
) -> Dict[str, Any]:
    price_block = _pick_price_block(raw, property_type)
    inferred_property_type = (
        property_type
        or price_block.get("type_bien")
        or raw.get("property_type")
        or raw.get("type_bien")
    )

    row = {
        "source": raw.get("source") or raw.get("site") or raw.get("scraper") or "source_inconnue",
        "localisation": raw.get("localisation"),
        "property_type": inferred_property_type,
        "prix_bas_m2": _coalesce_price(
            raw.get("prix_bas_m2"),
            raw.get("tranche_basse"),
            price_block.get("prix_m2_min"),
        ),
        "prix_moyen_m2": _coalesce_price(
            raw.get("prix_moyen_m2"),
            raw.get("prix_m2_moyen"),
            price_block.get("prix_m2_moyen"),
        ),
        "prix_haut_m2": _coalesce_price(
            raw.get("prix_haut_m2"),
            raw.get("tranche_haute"),
            price_block.get("prix_m2_max"),
        ),
        "url_source": raw.get("url_source"),
        "method_used": raw.get("method_used"),
        "error": raw.get("error"),
        "status": raw.get("status"),
    }
    row["status"] = _normalize_status(row)
    return row


def normalize_scraping_results(
    raw_results: Iterable[Dict[str, Any]],
    *,
    property_type: Optional[str] = None,
) -> List[Dict[str, Any]]:
    return [
        normalize_scraping_result(raw, property_type=property_type)
        for raw in raw_results
    ]


def _compute_global_average(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    valid_rows = [row for row in rows if row.get("status") == "ok"]

    def average(key: str) -> Optional[float]:
        values = [float(row[key]) for row in valid_rows if row.get(key) is not None]
        if not values:
            return None
        return round(sum(values) / len(values), 2)

    return {
        "source": "Moyenne globale",
        "localisation": None,
        "property_type": None,
        "prix_bas_m2": average("prix_bas_m2"),
        "prix_moyen_m2": average("prix_moyen_m2"),
        "prix_haut_m2": average("prix_haut_m2"),
        "status": "synthese",
        "url_source": None,
        "method_used": None,
        "error": None,
    }


def compute_scraping_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    summary_row = _compute_global_average(rows)
    succeeded_rows = [row for row in rows if row.get("status") == "ok"]
    return {
        "sources_total": len(rows),
        "sources_succeeded": len(succeeded_rows),
        "sources_failed": len([row for row in rows if row.get("status") == "error"]),
        "sources_empty": len([row for row in rows if row.get("status") == "empty"]),
        "sources_skipped": len([row for row in rows if row.get("status") == "skipped"]),
        "average_price_per_sqm_eur": summary_row.get("prix_moyen_m2"),
        "min_price_per_sqm_eur": min(
            (row["prix_moyen_m2"] for row in succeeded_rows if row.get("prix_moyen_m2") is not None),
            default=None,
        ),
        "max_price_per_sqm_eur": max(
            (row["prix_moyen_m2"] for row in succeeded_rows if row.get("prix_moyen_m2") is not None),
            default=None,
        ),
    }


def _style_header(ws, row_index: int) -> None:
    for column_index, (_, header) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=row_index, column=column_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data_row(ws, row_index: int, row: Dict[str, Any]) -> None:
    fill = None
    if row.get("status") == "synthese":
        fill = SUMMARY_FILL
    elif row.get("status") == "error":
        fill = ERROR_FILL
    elif row.get("status") == "ok":
        fill = SUCCESS_FILL

    for column_index, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=row_index, column=column_index, value=row.get(key))
        cell.font = VALUE_FONT
        cell.border = THIN_BORDER
        if fill is not None:
            cell.fill = fill
        apply_numeric_format(cell)

    ws.cell(row=row_index, column=1).font = BOLD_FONT
    if row.get("error"):
        ws.cell(row=row_index, column=len(EXPORT_COLUMNS)).font = RED_FONT


def _set_column_widths(ws) -> None:
    widths = {
        "A": 22,
        "B": 28,
        "C": 16,
        "D": 18,
        "E": 20,
        "F": 18,
        "G": 12,
        "H": 48,
        "I": 14,
        "J": 48,
    }
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width


def _write_intro(ws, rows_count: int) -> int:
    ws["A1"] = "Synthese des scrapers immobiliers"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    ws["A2"] = f"Nombre de sources traitees: {rows_count}"
    ws["A2"].font = LABEL_FONT
    return 4


def _add_chart(ws, data_start_row: int, data_end_row: int) -> None:
    if data_end_row < data_start_row:
        return

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Comparaison des prix au m² par source"
    chart.y_axis.title = "Prix en EUR/m²"
    chart.x_axis.title = "Sources"
    chart.height = 9
    chart.width = 20
    chart.varyColors = False

    data = Reference(
        ws,
        min_col=4,
        max_col=6,
        min_row=data_start_row,
        max_row=data_end_row,
    )
    categories = Reference(
        ws,
        min_col=1,
        min_row=data_start_row + 1,
        max_row=data_end_row,
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend.position = "r"

    for series in chart.series:
        series.graphicalProperties.solidFill = None

    ws.add_chart(chart, "L4")


def build_scraping_excel_export(
    raw_results: Iterable[Dict[str, Any]],
    *,
    property_type: Optional[str] = None,
) -> bytes:
    normalized_rows = normalize_scraping_results(raw_results, property_type=property_type)
    summary_row = _compute_global_average(normalized_rows)
    export_rows = [*normalized_rows, summary_row]

    wb = Workbook()
    ws = wb.active
    ws.title = "Scrapers"

    data_start_row = _write_intro(ws, len(normalized_rows))
    _style_header(ws, data_start_row)

    for index, row in enumerate(export_rows, start=data_start_row + 1):
        _style_data_row(ws, index, row)

    ws.freeze_panes = f"A{data_start_row + 1}"
    _set_column_widths(ws)
    ws.auto_filter.ref = f"A{data_start_row}:J{data_start_row + len(export_rows)}"

    chart_end_row = data_start_row + len(normalized_rows)
    if normalized_rows:
        _add_chart(ws, data_start_row, chart_end_row)

    for row in ws.iter_rows(
        min_row=data_start_row + 1,
        max_row=data_start_row + len(export_rows),
        min_col=8,
        max_col=10,
    ):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def save_scraping_excel_report(
    output_path: str,
    raw_results: Iterable[Dict[str, Any]],
    *,
    property_type: Optional[str] = None,
) -> str:
    excel_bytes = build_scraping_excel_export(raw_results, property_type=property_type)
    with open(output_path, "wb") as file_obj:
        file_obj.write(excel_bytes)
    return output_path
