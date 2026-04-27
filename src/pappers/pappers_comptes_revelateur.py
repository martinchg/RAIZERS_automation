import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font

_SRC_DIR = Path(__file__).parent.parent.resolve()
if str(_SRC_DIR) not in sys.path:
    sys.path.insert(0, str(_SRC_DIR))

from core.excel_utils import (
    BOLD_FONT,
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    SECTION_FONT,
    THIN_BORDER,
    VALUE_FONT,
    apply_numeric_format,
    format_display_value,
    to_number,
)
from extraction.extract_structured_runtime import call_llm_with_retry, parse_json_response
from core.llm_client import get_llm_client
from core.normalization import canonical_name
from pappers.pappers_comptes_flatten import flatten_pappers_comptes
from core.runtime_config import configure_environment

ROOT_DIR = _SRC_DIR.parent.resolve()
configure_environment(ROOT_DIR)

logger = logging.getLogger(__name__)


def _default_output_path(input_path: Path, suffix: str) -> Path:
    if input_path.suffix.lower() == ".json":
        return input_path.with_name(f"{input_path.stem}_{suffix}.json")
    return input_path.with_name(f"{input_path.name}_{suffix}.json")


def _default_output_excel_path(input_path: Path, suffix: str) -> Path:
    if input_path.suffix.lower() == ".json":
        return input_path.with_name(f"{input_path.stem}_{suffix}.xlsx")
    return input_path.with_name(f"{input_path.name}_{suffix}.xlsx")


def _ensure_compact_payload(payload: Dict[str, Any], source_path: Path) -> Dict[str, Any]:
    if isinstance(payload, dict) and isinstance(payload.get("entries"), list):
        return payload
    return flatten_pappers_comptes(payload, source_path=source_path, detailed=False)


def _select_entry(compact_payload: Dict[str, Any], year: Optional[str], entry_index: int) -> Dict[str, Any]:
    entries = compact_payload.get("entries") or []
    if year is not None:
        entries = [entry for entry in entries if str(entry.get("y")) == str(year)]
    if not entries:
        raise ValueError("Aucune entrée exploitable trouvée dans le fichier")
    if entry_index < 0 or entry_index >= len(entries):
        raise IndexError(f"entry_index hors borne: {entry_index} (entries={len(entries)})")
    return entries[entry_index]


def _build_mapping_prompt(
    entry: Dict[str, Any],
    *,
    title: str,
    max_lines_per_section: int,
    max_lines_compte_resultat: int,
) -> str:
    """Prompt étape 1 : le LLM décide les regroupements et les libellés.
    Il ne calcule AUCUNE valeur — seulement des listes de codes sources.
    """
    def _lines_block(rows: List[Dict]) -> str:
        return json.dumps(
            [{"code": r.get("c"), "libelle": r.get("l")} for r in rows if r.get("l")],
            ensure_ascii=False,
        )

    actif_block  = _lines_block(entry.get("actif")  or [])
    passif_block = _lines_block(entry.get("passif") or [])
    cr_block     = _lines_block(entry.get("compte_resultat") or [])

    return f"""Tu analyses des comptes annuels Pappers. Tu es expert en analyse financière.

TON UNIQUE RÔLE : décider comment regrouper les lignes comptables en un tableau révélateur lisible.
Tu ne calcules AUCUNE valeur. Tu n'écris AUCUN chiffre. Tu ne touches à AUCUN montant.

Chaque ligne source est identifiée par son `code`. Tu vas constituer des groupes en listant les codes.

═══ Lignes disponibles ═══

actif  : {actif_block}
passif : {passif_block}
compte_resultat : {cr_block}

═══ Règles ═══
- Retourne exactement 3 sections : `actif`, `passif`, `compte_resultat`.
- Chaque section est une liste d'objets `{{"label": "libellé lisible", "sources": ["CODE1", "CODE2"]}}`.
- `label` : libellé court et lisible (ex. "Immobilisations corporelles", "Capital social", "Résultat net").
- `sources` : liste des codes sources à agréger pour ce groupe. Ne jamais laisser vide [].
- Maximum {max_lines_per_section} groupes pour `actif` et `passif`.
- Maximum {max_lines_compte_resultat} groupes pour `compte_resultat`.
- Vise 8 à 10 groupes par section sauf si l'information l'exige.
- Un code ne peut apparaître que dans UN SEUL groupe (pas de doublon).
- Tu peux regrouper des lignes proches uniquement si c'est fidèle et non ambigu.
- RÈGLE CRITIQUE — sous-totaux : un groupe ne peut jamais mélanger un sous-total ou total avec les lignes de détail qu'il agrège déjà. Soit tu utilises le total seul (préféré), soit tu utilises les lignes de détail seules — jamais les deux ensemble dans le même groupe. Exemple interdit : [TOTAL_I, TOTAL_II, TOTAL_GENERAL] car TOTAL_GENERAL inclut déjà TOTAL_I et TOTAL_II.
- Pour un groupe de total général (Total Actif, Total Passif, Total Général), utilise uniquement le code du total global — jamais ses sous-totaux ni ses lignes de détail.
- Si un total général explicite existe, crée un groupe pour lui en dernière position.
- Si une section est vide → retourne `[]`.
- INTERDIT : écrire un nombre, inventer un code, modifier un code existant.

═══ Format de sortie (JSON strict, aucun texte autour) ═══
{{
  "actif": [
    {{"label": "Immobilisations corporelles", "sources": ["BB", "BJ"]}},
    {{"label": "Total Actif", "sources": ["EE"]}}
  ],
  "passif": [...],
  "compte_resultat": [...]
}}
"""


def _resolve_pappers_mapping(
    mapping: Dict[str, Any],
    entry: Dict[str, Any],
) -> Dict[str, List[Dict[str, Any]]]:
    """Applique le mapping LLM sur les vraies valeurs de l'entrée Pappers aplatie.

    Lookup par code (fiable) avec fallback label normalisé.
    """
    import unicodedata

    def _norm(s: str) -> str:
        nfd = unicodedata.normalize("NFD", s.lower().strip())
        return "".join(c for c in nfd if unicodedata.category(c) != "Mn")

    SECTION_KEYS = ("actif", "passif", "compte_resultat")

    def _build_lookup(rows: List[Dict]) -> tuple:
        by_code: Dict[str, tuple] = {}
        by_norm: Dict[str, tuple] = {}
        for r in rows:
            code  = str(r.get("c") or "").strip()
            label = str(r.get("l") or "").strip()
            val   = (r.get("n"), r.get("n1"))
            if code:
                by_code[code] = val
            if label:
                by_norm[_norm(label)] = val
        return by_code, by_norm

    def _get(source: str, by_code: Dict, by_norm: Dict) -> tuple:
        if source in by_code:
            return by_code[source]
        if _norm(source) in by_norm:
            return by_norm[_norm(source)]
        return (None, None)

    result: Dict[str, List] = {}
    for section in SECTION_KEYS:
        rows = entry.get(section) or []
        by_code, by_norm = _build_lookup(rows)
        groups = mapping.get(section) or []

        if not groups:
            result[section] = []
            continue

        section_out = []
        for group in groups:
            label   = str(group.get("label") or "").strip()
            sources = group.get("sources") or []
            if not label:
                continue

            n_vals, n1_vals = [], []
            for src in sources:
                n_val, n1_val = _get(src, by_code, by_norm)
                if n_val  is not None: n_vals.append(n_val)
                if n1_val is not None: n1_vals.append(n1_val)

            section_out.append({
                "label": label,
                "n":  sum(n_vals)  if n_vals  else None,
                "n1": sum(n1_vals) if n1_vals else None,
            })
        result[section] = section_out

    return result



def _coerce_number_or_none(value: Any) -> Optional[float]:
    number = to_number(value)
    if number is None:
        return None
    return number


def _clean_section_rows(rows: Any, max_lines_per_section: int) -> List[Dict[str, Any]]:
    if not isinstance(rows, list):
        return []

    cleaned: List[Dict[str, Any]] = []
    seen_labels: set = set()
    seen_values: set = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        label = str(row.get("label") or row.get("libelle") or "").strip()
        if not label:
            continue
        if label.lower() in seen_labels:
            continue
        n  = _coerce_number_or_none(row.get("n"))
        n1 = _coerce_number_or_none(row.get("n1"))
        val_key = (n, n1)
        if n is not None and val_key in seen_values:
            continue
        seen_labels.add(label.lower())
        if n is not None:
            seen_values.add(val_key)
        cleaned.append({"label": label, "n": n, "n1": n1})
        if len(cleaned) >= max_lines_per_section:
            break
    return cleaned


def _clean_llm_output(
    raw: Dict[str, Any],
    *,
    max_lines_per_section: int,
    max_lines_compte_resultat: int,
) -> Dict[str, List[Dict[str, Any]]]:
    return {
        "actif": _clean_section_rows(raw.get("actif"), max_lines_per_section),
        "passif": _clean_section_rows(raw.get("passif"), max_lines_per_section),
        "compte_resultat": _clean_section_rows(raw.get("compte_resultat"), max_lines_compte_resultat),
    }


def _format_period_label(date_value: Optional[str], fallback: str) -> str:
    if not date_value:
        return fallback
    try:
        return f"Au {datetime.strptime(date_value, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    except ValueError:
        return f"Au {date_value}"


def _row_font_for_label(label: str) -> Font:
    canon = canonical_name(label)
    highlight_tokens = (
        "total",
        "capitaux propres",
        "resultat",
        "benefice ou perte",
    )
    if any(token in canon for token in highlight_tokens):
        return BOLD_FONT
    return VALUE_FONT


def _write_excel_section(
    ws,
    start_row: int,
    title: str,
    rows: List[Dict[str, Any]],
    *,
    date_n: Optional[str],
    date_n1: Optional[str],
) -> int:
    title_cell = ws.cell(row=start_row, column=2, value=title)
    title_cell.font = SECTION_FONT
    start_row += 1

    for col, header in (
        (2, "En €"),
        (3, _format_period_label(date_n, "Exercice N")),
        (4, _format_period_label(date_n1, "Exercice N-1")),
    ):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    start_row += 1

    if not rows:
        for col in range(2, 5):
            cell = ws.cell(row=start_row, column=col, value="" if col != 2 else "Aucune donnée")
            cell.border = THIN_BORDER
            cell.font = VALUE_FONT
        return start_row + 2

    for row in rows:
        row_font = _row_font_for_label(str(row.get("label") or ""))
        label_cell = ws.cell(row=start_row, column=2, value=row.get("label"))
        label_cell.font = row_font
        label_cell.border = THIN_BORDER

        for col, key in ((3, "n"), (4, "n1")):
            cell = ws.cell(row=start_row, column=col)
            cell.border = THIN_BORDER
            cell.font = row_font
            value = row.get(key)
            number = to_number(value)
            if number is not None:
                cell.value = number
                apply_numeric_format(cell, number)
            elif value is not None:
                cell.value = format_display_value(value)
        start_row += 1

    return start_row + 1


def write_revealing_excel(
    output_path: Path,
    table_data: Dict[str, Any],
    *,
    title: str,
    date_cloture: Optional[str],
    date_cloture_n1: Optional[str],
    type_comptes: Optional[str],
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tableau révélateur"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    row = 1
    title_cell = ws.cell(row=row, column=2, value=title)
    title_cell.font = Font(bold=True, size=13, color="2F5496")
    row += 1

    meta_bits = []
    if date_cloture:
        meta_bits.append(f"Clôture: {date_cloture}")
    if type_comptes:
        meta_bits.append(f"Type: {type_comptes}")
    if meta_bits:
        meta_cell = ws.cell(row=row, column=2, value=" | ".join(meta_bits))
        meta_cell.font = BOLD_FONT
        row += 2

    row = _write_excel_section(
        ws,
        row,
        "Actif",
        table_data.get("actif") or [],
        date_n=date_cloture,
        date_n1=date_cloture_n1,
    )
    row = _write_excel_section(
        ws,
        row,
        "Passif",
        table_data.get("passif") or [],
        date_n=date_cloture,
        date_n1=date_cloture_n1,
    )
    _write_excel_section(
        ws,
        row,
        "Compte de résultat",
        table_data.get("compte_resultat") or [],
        date_n=date_cloture,
        date_n1=date_cloture_n1,
    )

    wb.save(str(output_path))
    return output_path


def run(
    input_path: Path,
    *,
    year: Optional[str] = None,
    entry_index: int = 0,
    title: Optional[str] = None,
    max_lines_per_section: int = 15,
    max_lines_compte_resultat: int = 15,
    prompt_only: bool = False,
    output_json: Optional[Path] = None,
    output_excel: Optional[Path] = None,
) -> Dict[str, Any]:
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    compact_payload = _ensure_compact_payload(payload, input_path)
    selected_entry = _select_entry(compact_payload, year, entry_index)
    resolved_title = title or input_path.stem
    prompt = _build_mapping_prompt(
        selected_entry,
        title=resolved_title,
        max_lines_per_section=max_lines_per_section,
        max_lines_compte_resultat=max_lines_compte_resultat,
    )

    if prompt_only:
        return {
            "prompt": prompt,
            "selected_entry": selected_entry,
        }

    llm_client = get_llm_client(
        model_override={"openai": "gpt-4o-mini", "gemini": "gemini-2.5-flash-lite"},
        preferred_provider="gemini",
    )
    raw_response = call_llm_with_retry(llm_client["text_call"], prompt)
    mapping = parse_json_response(raw_response)
    resolved = _resolve_pappers_mapping(mapping, selected_entry)
    cleaned = _clean_llm_output(
        resolved,
        max_lines_per_section=max_lines_per_section,
        max_lines_compte_resultat=max_lines_compte_resultat,
    )

    result_payload = {
        "title": resolved_title,
        "source_file": str(input_path),
        "model": llm_client["model"],
        "exercise": {
            "year": selected_entry.get("y"),
            "date_cloture": selected_entry.get("cloture"),
            "date_cloture_n1": selected_entry.get("cloture_n1"),
            "type_comptes": selected_entry.get("type"),
        },
        "tableau_revelateur": cleaned,
    }

    if output_json:
        output_json.write_text(json.dumps(result_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    if output_excel:
        write_revealing_excel(
            output_excel,
            cleaned,
            title=resolved_title,
            date_cloture=selected_entry.get("cloture"),
            date_cloture_n1=selected_entry.get("cloture_n1"),
            type_comptes=selected_entry.get("type"),
        )

    return result_payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Génère un tableau financier révélateur via LLM à partir de Pappers /entreprise/comptes")
    parser.add_argument("--input", "-i", required=True, help="JSON brut Pappers ou JSON compact flatten")
    parser.add_argument("--year", default=None, help="Année à sélectionner si plusieurs exercices sont présents")
    parser.add_argument("--entry-index", type=int, default=0, help="Index de l'entrée parmi les exercices filtrés")
    parser.add_argument("--title", default=None, help="Titre affiché dans les sorties JSON/Excel")
    parser.add_argument("--max-lines-per-section", type=int, default=15, help="Nombre maximal de lignes par section")
    parser.add_argument(
        "--max-lines-compte-resultat",
        type=int,
        default=15,
        help="Nombre maximal de lignes pour le compte de résultat",
    )
    parser.add_argument("--prompt-only", action="store_true", help="N'appelle pas le LLM, renvoie seulement le prompt et l'entrée sélectionnée")
    parser.add_argument("--output-json", default=None, help="Chemin du JSON de sortie")
    parser.add_argument("--output-excel", default=None, help="Chemin de l'Excel de sortie")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    output_json = Path(args.output_json).resolve() if args.output_json else _default_output_path(input_path, "revealing")
    output_excel = Path(args.output_excel).resolve() if args.output_excel else _default_output_excel_path(input_path, "revealing")

    payload = run(
        input_path,
        year=args.year,
        entry_index=args.entry_index,
        title=args.title,
        max_lines_per_section=args.max_lines_per_section,
        max_lines_compte_resultat=args.max_lines_compte_resultat,
        prompt_only=args.prompt_only,
        output_json=None if args.prompt_only else output_json,
        output_excel=None if args.prompt_only else output_excel,
    )

    if args.prompt_only:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        return

    print(output_json)
    print(output_excel)


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s — %(levelname)s — %(message)s",
    )
    main()
