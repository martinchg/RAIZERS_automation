from __future__ import annotations

import calendar
import math
import re
import time
import unicodedata
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional
from urllib.parse import quote

import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, ValidationError, field_validator

from core.excel_utils import (
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    HYPERLINK_FONT,
    THIN_BORDER,
    VALUE_FONT,
    apply_numeric_format,
)
from immo_scoring import (
    ComparableScorer,
    PropertyType,
    SubjectProperty,
    normalize_micro_location,
)


# -----------------------------------------------------------------------------
# Configuration
# -----------------------------------------------------------------------------

GEOCODER_URL = "https://data.geopf.fr/geocodage/search"
DVF_API_BASE_URL = "https://apidf-preprod.cerema.fr"
DVF_API_TOKEN = None
DEFAULT_TIMEOUT_SECONDS = 30
USER_AGENT = "comparateur-immo-streamlit/0.1"
DEFAULT_RETRY_COUNT = 3
DEFAULT_BACKOFF_SECONDS = 0.8
MAX_RETAINED_COMPARABLES = 10
DVF_MONTH_WINDOW = 20
DVF_PAGE_SIZE = 100
DEFAULT_SEARCH_BBOX_DELTA = 0.002
DEFAULT_SEARCH_RADIUS_M = int(round(DEFAULT_SEARCH_BBOX_DELTA * 111320))


# -----------------------------------------------------------------------------
# Input / output models
# -----------------------------------------------------------------------------


class ComparableRequest(BaseModel):
    address: str = Field(..., min_length=5)
    property_type: PropertyType
    living_area_sqm: float = Field(..., gt=0)
    rooms: int = Field(..., ge=1, le=20)
    land_area_sqm: Optional[float] = Field(None, ge=0)
    search_radius_m: int = Field(DEFAULT_SEARCH_RADIUS_M, ge=50, le=5000)
    api_min_year: Optional[int] = Field(None, ge=2000, le=2100)

    @field_validator("land_area_sqm")
    @classmethod
    def normalize_land_area(cls, value: Optional[float]) -> Optional[float]:
        if value is None:
            return None
        return round(float(value), 2)

    @field_validator("api_min_year")
    @classmethod
    def validate_api_min_year(cls, value: Optional[int]) -> Optional[int]:
        if value is None:
            return None
        minimum_year = 2000
        maximum_year = subtract_months(date.today(), DVF_MONTH_WINDOW).year
        if value < minimum_year or value > maximum_year:
            raise ValueError(
                f"L'année min API doit être comprise entre {minimum_year} et {maximum_year}"
            )
        return value


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------


def haversine_distance_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    r = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = (
        math.sin(dphi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    )
    return 2 * r * math.atan2(math.sqrt(a), math.sqrt(1 - a))



def median(values: List[float]) -> Optional[float]:
    ordered = sorted(v for v in values if v is not None)
    if not ordered:
        return None
    n = len(ordered)
    mid = n // 2
    if n % 2 == 1:
        return ordered[mid]
    return (ordered[mid - 1] + ordered[mid]) / 2



def format_french_number(value: Any) -> Any:
    if value is None or isinstance(value, bool) or not isinstance(value, (int, float)):
        return value
    rounded = round(float(value), 2)
    if math.isclose(rounded, round(rounded), abs_tol=1e-9):
        return f"{int(round(rounded)):,}".replace(",", " ")
    text = f"{rounded:,.2f}".replace(",", " ").replace(".", ",")
    if text.endswith("0"):
        text = text[:-1]
    if text.endswith(",0"):
        text = text[:-2]
    return text


def format_subject_value(key: str, value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if key in {
        "latitude",
        "longitude",
        "living_area_sqm",
        "land_area_sqm",
    }:
        return format_french_number(value)
    return value


def build_subject_display_rows(subject: Dict[str, Any]) -> List[Dict[str, Any]]:
    labels = {
        "normalized_address": "Adresse",
        "property_type": "Type de bien",
        "living_area_sqm": "Surface habitable (m²)",
        "rooms": "Nombre de pièces",
        "land_area_sqm": "Surface terrain (m²)",
        "city": "Ville",
        "postcode": "Code postal",
        "latitude": "Latitude",
        "longitude": "Longitude",
    }
    ordered_keys = [
        "normalized_address",
        "property_type",
        "living_area_sqm",
        "rooms",
        "land_area_sqm",
        "city",
        "postcode",
        "latitude",
        "longitude",
    ]

    rows: List[Dict[str, Any]] = []
    for key in ordered_keys:
        value = subject.get(key)
        if value is None:
            continue
        rows.append({
            "Champ": labels.get(key, key),
            "Valeur": format_subject_value(key, value),
        })
    return rows


def append_comparables_summary_row(
    comparables: List[Dict[str, Any]],
    statistics: Dict[str, Any],
) -> List[Dict[str, Any]]:
    if not comparables:
        return comparables

    rows = [dict(row) for row in comparables]
    summary_row = {key: "" for key in rows[0].keys()}
    summary_row["Retenu"] = "Synthèse"
    summary_row["Adresse"] = "Moyenne de tous les comparables"
    summary_row["Prix de vente"] = format_french_number(statistics.get("average_total_price_eur"))
    summary_row["Prix par m²"] = format_french_number(statistics.get("average_price_per_sqm_eur"))

    rows.append(summary_row)
    return rows


def compute_price_tranche_averages(values: List[float]) -> Dict[str, Optional[float]]:
    cleaned = sorted(float(value) for value in values if value is not None)
    if not cleaned:
        return {
            "average_price_per_sqm_low_band_eur": None,
            "average_price_per_sqm_mid_band_eur": None,
            "average_price_per_sqm_high_band_eur": None,
        }

    chunk_size, remainder = divmod(len(cleaned), 3)
    sizes = [chunk_size + (1 if index < remainder else 0) for index in range(3)]

    tranches: List[List[float]] = []
    start = 0
    for size in sizes:
        end = start + size
        tranches.append(cleaned[start:end])
        start = end

    def _average(items: List[float]) -> Optional[float]:
        if not items:
            return None
        return round(sum(items) / len(items), 2)

    return {
        "average_price_per_sqm_low_band_eur": _average(tranches[0]),
        "average_price_per_sqm_mid_band_eur": _average(tranches[1]),
        "average_price_per_sqm_high_band_eur": _average(tranches[2]),
    }


_STATS_LABELS_FR: Dict[str, str] = {
    "search_radius_m_used": "Rayon de recherche (m)",
    "api_min_year_used": "Année minimale",
    "comparables_found": "Comparables trouvés",
    "comparables_retained": "Comparables retenus",
    "average_total_price_eur": "Prix de vente moyen (€)",
    "min_price_per_sqm_eur": "Prix/m² minimum (€)",
    "median_price_per_sqm_eur": "Prix/m² médian (€)",
    "max_price_per_sqm_eur": "Prix/m² maximum (€)",
    "average_price_per_sqm_eur": "Prix/m² moyen (€)",
    "average_price_per_sqm_low_band_eur": "Prix/m² moyen tranche basse (€)",
    "average_price_per_sqm_mid_band_eur": "Prix/m² moyen tranche médiane (€)",
    "average_price_per_sqm_high_band_eur": "Prix/m² moyen tranche haute (€)",
}

_HIDDEN_COLS = {"_latitude", "_longitude"}


def _build_maps_url(lat: Optional[float], lon: Optional[float], address: Optional[str]) -> Optional[str]:
    if lat is not None and lon is not None:
        return f"https://www.google.com/maps?q={lat},{lon}"
    if address:
        return f"https://www.google.com/maps/search/{quote(str(address))}"
    return None


def build_immo_excel_export(result: Dict[str, Any]) -> bytes:
    output = BytesIO()
    comparables_rows = append_comparables_summary_row(
        result.get("comparables", []),
        result.get("statistics", {}),
    )
    statistics = result.get("statistics", {}) or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparables"

    # ---- Determine display columns (exclude hidden metadata) ----
    display_cols: List[str] = []
    if comparables_rows:
        display_cols = [col for col in comparables_rows[0].keys() if col not in _HIDDEN_COLS]

    adresse_col_idx = (display_cols.index("Adresse") + 1) if "Adresse" in display_cols else None

    # ---- Write comparables header ----
    for col_idx, header in enumerate(display_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # ---- Write comparables rows ----
    for row_offset, row_data in enumerate(comparables_rows):
        excel_row = row_offset + 2
        lat = row_data.get("_latitude")
        lon = row_data.get("_longitude")
        address = row_data.get("Adresse")
        maps_url = _build_maps_url(lat, lon, address)

        for col_idx, col_name in enumerate(display_cols, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=row_data.get(col_name))
            cell.font = VALUE_FONT
            cell.border = THIN_BORDER
            apply_numeric_format(cell)

            if col_idx == adresse_col_idx and maps_url and address:
                cell.hyperlink = maps_url
                cell.font = HYPERLINK_FONT

    # ---- Statistics section (French labels, embedded below comparables) ----
    stats_start_row = len(comparables_rows) + 4

    header_cell = ws.cell(row=stats_start_row, column=1, value="Statistiques")
    header_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_cell.fill = HEADER_FILL
    header_cell.alignment = HEADER_ALIGNMENT
    header_cell.border = THIN_BORDER

    label_header = ws.cell(row=stats_start_row, column=2, value="Valeur")
    label_header.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    label_header.fill = HEADER_FILL
    label_header.alignment = HEADER_ALIGNMENT
    label_header.border = THIN_BORDER

    for stat_offset, (key, value) in enumerate(statistics.items()):
        stat_row = stats_start_row + 1 + stat_offset
        label_cell = ws.cell(row=stat_row, column=1, value=_STATS_LABELS_FR.get(key, key))
        label_cell.font = VALUE_FONT
        label_cell.border = THIN_BORDER

        val_cell = ws.cell(row=stat_row, column=2, value=format_subject_value(key, value))
        val_cell.font = VALUE_FONT
        val_cell.border = THIN_BORDER
        apply_numeric_format(val_cell)

    # ---- Column widths ----
    for col_idx in range(1, len(display_cols) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 40)

    ws.freeze_panes = "A2"

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _style_immo_export_sheet(worksheet: Any) -> None:
    worksheet.freeze_panes = "A2"

    for row_idx, row in enumerate(worksheet.iter_rows(), start=1):
        for cell in row:
            cell.border = THIN_BORDER
            if row_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
            else:
                cell.font = VALUE_FONT
                apply_numeric_format(cell)

    for column_idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[get_column_letter(column_idx)].width = min(
            max(max_length + 2, 12),
            40,
        )


def try_parse_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    text = str(value)[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def subtract_months(value: date, months: int) -> date:
    year_delta, month_index = divmod(value.month - 1 - months, 12)
    target_year = value.year + year_delta
    target_month = month_index + 1
    target_day = min(value.day, calendar.monthrange(target_year, target_month)[1])
    return date(target_year, target_month, target_day)


def http_get_json(
    url: str,
    *,
    params: Optional[Dict[str, Any]] = None,
    headers: Optional[Dict[str, str]] = None,
    timeout: float = DEFAULT_TIMEOUT_SECONDS,
    retries: int = DEFAULT_RETRY_COUNT,
    backoff_seconds: float = DEFAULT_BACKOFF_SECONDS,
) -> Any:
    last_exception: Optional[Exception] = None

    for attempt in range(retries + 1):
        try:
            response = requests.get(
                url,
                params=params,
                headers=headers,
                timeout=(10, timeout),
            )
            response.raise_for_status()
            return response.json()
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as exc:
            last_exception = exc
            if attempt >= retries:
                raise

            # Exponential backoff to absorb temporary network glitches.
            time.sleep(backoff_seconds * (2**attempt))

    if last_exception:
        raise last_exception

    raise RuntimeError("Erreur réseau inattendue")


def reverse_geocode(longitude: float, latitude: float) -> Dict[str, Any]:
    headers = {"User-Agent": USER_AGENT}
    params = {
        "lon": longitude,
        "lat": latitude,
    }

    payload = http_get_json(
        "https://data.geopf.fr/geocodage/reverse",
        params=params,
        headers=headers,
    )

    features = payload.get("features", [])
    if not features:
        return {}

    props = features[0].get("properties", {})
    return {
        "label": props.get("label"),
        "city": props.get("city"),
        "postcode": props.get("postcode"),
        "street": props.get("street"),
        "housenumber": props.get("housenumber"),
    }


# -----------------------------------------------------------------------------
# External clients
# -----------------------------------------------------------------------------


class GeocoderClient:
    def __init__(self, base_url: str = GEOCODER_URL) -> None:
        self.base_url = base_url

    def geocode(self, address: str) -> Dict[str, Any]:
        params = {"q": address, "limit": 1}
        headers = {"User-Agent": USER_AGENT}
        payload = http_get_json(
            self.base_url,
            params=params,
            headers=headers,
        )

        features = payload.get("features", [])
        if not features:
            raise ValueError("Adresse non trouvée")

        feature = features[0]
        props = feature.get("properties", {})
        coordinates = feature.get("geometry", {}).get("coordinates", [])
        if len(coordinates) != 2:
            raise ValueError("Géocodage incomplet")

        return {
            "normalized_address": props.get("label") or address,
            "street_name": props.get("street") or props.get("name"),
            "city": props.get("city"),
            "postcode": props.get("postcode"),
            "longitude": float(coordinates[0]),
            "latitude": float(coordinates[1]),
        }


def _normalize_address_text(value: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    ascii_text = ascii_text.lower()
    ascii_text = re.sub(r"[^a-z0-9]+", " ", ascii_text)
    return " ".join(ascii_text.split())


def _build_address_suggestion(item: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if "properties" in item:
        props = item.get("properties", {})
        coords = item.get("geometry", {}).get("coordinates", [])
        longitude = coords[0] if len(coords) == 2 else None
        latitude = coords[1] if len(coords) == 2 else None
    else:
        props = item
        longitude = item.get("x") or item.get("lon") or item.get("longitude")
        latitude = item.get("y") or item.get("lat") or item.get("latitude")

    label = props.get("fulltext") or props.get("label") or props.get("name") or ""
    if not label:
        return None

    return {
        "label": label,
        "city": props.get("city"),
        "postcode": props.get("zipcode") or props.get("postcode"),
        "street": props.get("street"),
        "longitude": float(longitude) if longitude is not None else None,
        "latitude": float(latitude) if latitude is not None else None,
    }


def get_address_suggestions(query: str, limit: int = 6) -> List[Dict[str, Any]]:
    if not query or len(query.strip()) < 3:
        return []

    query = query.strip()
    limit = max(1, min(limit, 15))
    headers = {"User-Agent": USER_AGENT}
    params = {
        "text": query,
        "type": "StreetAddress",
        "maximumResponses": limit,
    }

    payload = http_get_json(
        "https://data.geopf.fr/geocodage/completion",
        params=params,
        headers=headers,
    )

    raw_items = payload.get("results") or payload.get("features") or []
    suggestions: List[Dict[str, Any]] = []
    seen_labels: set[str] = set()

    for item in raw_items:
        suggestion = _build_address_suggestion(item)
        if not suggestion:
            continue

        normalized_label = _normalize_address_text(suggestion["label"])
        if normalized_label in seen_labels:
            continue

        seen_labels.add(normalized_label)
        suggestions.append(suggestion)

    if suggestions:
        return suggestions[:limit]

    # Fallback: the search endpoint includes fuzzy autocomplete and returns GeoJSON.
    fallback_payload = http_get_json(
        GEOCODER_URL,
        params={
            "q": query,
            "limit": limit,
            "type": "StreetAddress",
            "autocomplete": 1,
        },
        headers=headers,
    )

    for item in fallback_payload.get("features", []):
        suggestion = _build_address_suggestion(item)
        if not suggestion:
            continue

        normalized_label = _normalize_address_text(suggestion["label"])
        if normalized_label in seen_labels:
            continue

        seen_labels.add(normalized_label)
        suggestions.append(suggestion)

    return suggestions[:limit]

def _meters_to_lat_delta(meters: float) -> float:
    return meters / 111320


def _meters_to_lon_delta(meters: float, latitude: float) -> float:
    latitude_factor = max(abs(math.cos(math.radians(latitude))), 0.1)
    return meters / (111320 * latitude_factor)


def _bbox_around(
    longitude: float,
    latitude: float,
    radius_m: float,
) -> str:
    lat_delta = _meters_to_lat_delta(radius_m)
    lon_delta = _meters_to_lon_delta(radius_m, latitude)

    lon_min = longitude - lon_delta
    lon_max = longitude + lon_delta
    lat_min = latitude - lat_delta
    lat_max = latitude + lat_delta

    return f"{lon_min},{lat_min},{lon_max},{lat_max}"


def _geometry_center(geometry: Dict[str, Any]) -> tuple[Optional[float], Optional[float]]:
    if not geometry:
        return None, None

    gtype = geometry.get("type")
    coords = geometry.get("coordinates")

    if gtype == "Point" and isinstance(coords, list) and len(coords) >= 2:
        return float(coords[0]), float(coords[1])

    if gtype == "Polygon" and coords and coords[0]:
        ring = coords[0]
        xs = [p[0] for p in ring if isinstance(p, list) and len(p) >= 2]
        ys = [p[1] for p in ring if isinstance(p, list) and len(p) >= 2]
        if xs and ys:
            return sum(xs) / len(xs), sum(ys) / len(ys)

    if gtype == "MultiPolygon" and coords and coords[0] and coords[0][0]:
        ring = coords[0][0]
        xs = [p[0] for p in ring if isinstance(p, list) and len(p) >= 2]
        ys = [p[1] for p in ring if isinstance(p, list) and len(p) >= 2]
        if xs and ys:
            return sum(xs) / len(xs), sum(ys) / len(ys)

    return None, None


class DVFClient:
    def __init__(self, base_url: str = DVF_API_BASE_URL, token: Optional[str] = DVF_API_TOKEN) -> None:
        self.base_url = base_url.rstrip("/")
        self.token = token

    def search_transactions(
        self,
        *,
        latitude: float,
        longitude: float,
        property_type: PropertyType,
        living_area_sqm: float,
        rooms: int,
        land_area_sqm: Optional[float],
        valuation_date: date,
        search_radius_m: int = DEFAULT_SEARCH_RADIUS_M,
        api_min_year: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        headers = {"Content-Type": "application/json"}
        if self.token:
            headers["Authorization"] = "Token " + self.token

        default_start_date = subtract_months(valuation_date, DVF_MONTH_WINDOW)
        requested_start_date = (
            date(int(api_min_year), 1, 1)
            if api_min_year is not None
            else default_start_date
        )
        effective_api_min_year = requested_start_date.year

        params: Dict[str, Any] = {
            "anneemut_min": str(effective_api_min_year),
            "anneemut_max": str(valuation_date.year),
            "fields": "all",
            "page_size": DVF_PAGE_SIZE,
            "format": "json",
        }

        records = self._fetch_geomutations_page_set(
            headers=headers,
            params={
                **params,
                "in_bbox": _bbox_around(longitude, latitude, search_radius_m),
            },
        )

        filtered_records: List[Dict[str, Any]] = []
        for record in records:
            props = record.get("properties", record)
            sale_date = try_parse_date(props.get("datemut") or props.get("date_mutation"))
            if sale_date is None:
                continue
            if requested_start_date <= sale_date <= valuation_date:
                filtered_records.append(record)

        filtered_records.sort(
            key=lambda item: try_parse_date(
                (item.get("properties", item)).get("datemut")
                or (item.get("properties", item)).get("date_mutation")
            )
            or date.min,
            reverse=True,
        )
        return filtered_records

    def _fetch_geomutations_page_set(
        self,
        *,
        headers: Dict[str, str],
        params: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        records: List[Dict[str, Any]] = []
        expected_count: Optional[int] = None
        page = 1

        while True:
            payload = http_get_json(
                f"{self.base_url}/dvf_opendata/geomutations/",
                params={**params, "page": page},
                headers=headers,
                timeout=60,
            )

            page_records: List[Dict[str, Any]] = []
            has_next_page = False
            if isinstance(payload, dict):
                if "features" in payload and isinstance(payload["features"], list):
                    page_records = payload["features"]
                elif "results" in payload and isinstance(payload["results"], list):
                    page_records = payload["results"]

                count = payload.get("count")
                if isinstance(count, int):
                    expected_count = count
                has_next_page = bool(payload.get("next"))
            elif isinstance(payload, list):
                page_records = payload

            if not page_records:
                break

            records.extend(page_records)

            if expected_count is not None and len(records) >= expected_count:
                break
            if has_next_page or (expected_count is not None and len(records) < expected_count):
                page += 1
                continue
            if len(page_records) < DVF_PAGE_SIZE:
                break

            page += 1

        return records

    @staticmethod
    def _record_key(record: Dict[str, Any]) -> str:
        props = record.get("properties", record)
        explicit_id = props.get("idmutation") or props.get("idmutinvar") or props.get("idnatmut")
        if explicit_id:
            return str(explicit_id)

        longitude, latitude = _geometry_center(record.get("geometry", {}))
        fallback_parts = [
            str(props.get("datemut") or props.get("date_mutation") or ""),
            str(props.get("valeurfonc") or props.get("valeur_fonciere") or ""),
            str(props.get("sbati") or props.get("surface_reelle_bati") or ""),
            str(props.get("sterr") or props.get("surface_terrain") or ""),
            str(longitude or ""),
            str(latitude or ""),
        ]
        return "|".join(fallback_parts)


class ComparablePipeline:
    def __init__(self, geocoder: GeocoderClient, dvf_client: DVFClient, scorer: ComparableScorer) -> None:
        self.geocoder = geocoder
        self.dvf_client = dvf_client
        self.scorer = scorer

    def run(self, payload: ComparableRequest) -> Dict[str, Any]:
        valuation_date = date.today()
        geo = self.geocoder.geocode(payload.address)

        subject = SubjectProperty(
            normalized_address=geo["normalized_address"],
            latitude=geo["latitude"],
            longitude=geo["longitude"],
            city=geo.get("city"),
            postcode=geo.get("postcode"),
            street_name=geo.get("street_name"),
            property_type=payload.property_type,
            living_area_sqm=payload.living_area_sqm,
            rooms=payload.rooms,
            land_area_sqm=payload.land_area_sqm,
        )

        raw_records = self.dvf_client.search_transactions(
            latitude=subject.latitude,
            longitude=subject.longitude,
            property_type=subject.property_type,
            living_area_sqm=subject.living_area_sqm,
            rooms=subject.rooms,
            land_area_sqm=subject.land_area_sqm,
            valuation_date=valuation_date,
            search_radius_m=payload.search_radius_m,
            api_min_year=payload.api_min_year,
        )

        normalized_comparables: List[Dict[str, Any]] = []
        for raw in raw_records:
            comp = self._normalize_record(raw, subject)
            if comp is None:
                continue
            normalized_comparables.append(comp)

        deduplicated_comparables = self._deduplicate_recent_same_address_sales(normalized_comparables)

        eligible_comparables: List[Dict[str, Any]] = []
        for comp in deduplicated_comparables:
            distance_m = comp.get("distance_m")
            if distance_m is not None and distance_m > payload.search_radius_m:
                continue
            comp["similarity_score"] = self.scorer.score(
                subject,
                comp,
                reference_date=valuation_date,
            )
            eligible_comparables.append(comp)

        eligible_prices = [
            comp["price_per_sqm_eur"]
            for comp in eligible_comparables
            if comp.get("price_per_sqm_eur") is not None
        ]

        sorted_comparables = self._sort_comparables(eligible_comparables)
        retained_comparables = self._select_comparables(subject, sorted_comparables)
        retained_ids = {id(comp) for comp in retained_comparables}
        all_total_prices = [
            comp["total_price_eur"]
            for comp in sorted_comparables
            if comp.get("total_price_eur") is not None
        ]
        all_prices = [
            comp["price_per_sqm_eur"]
            for comp in sorted_comparables
            if comp.get("price_per_sqm_eur") is not None
        ]
        returned_comparables = [
            self._format_comparable_for_display(comp, retained=id(comp) in retained_ids)
            for comp in sorted_comparables
        ]
        median_price = median(all_prices)
        price_tranches = compute_price_tranche_averages(all_prices)

        return {
            "subject": subject.model_dump(),
            "comparables": returned_comparables,
            "statistics": {
                "search_radius_m_used": payload.search_radius_m,
                "api_min_year_used": payload.api_min_year,
                "comparables_found": len(eligible_comparables),
                "comparables_retained": len(retained_comparables),
                "average_total_price_eur": round(sum(all_total_prices) / len(all_total_prices), 2) if all_total_prices else None,
                "min_price_per_sqm_eur": min(all_prices) if all_prices else None,
                "median_price_per_sqm_eur": round(median_price, 2) if median_price is not None else None,
                "max_price_per_sqm_eur": max(all_prices) if all_prices else None,
                "average_price_per_sqm_eur": round(sum(all_prices) / len(all_prices), 2) if all_prices else None,
                **price_tranches,
            },
        }

    def _normalize_record(self, raw: Dict[str, Any], subject: SubjectProperty) -> Optional[Dict[str, Any]]:
        props = raw.get("properties", raw)
        property_type = self._normalize_property_type(props)
        living_area = self._to_float(props.get("sbati") or props.get("surface_reelle_bati"))
        land_area = self._to_float(props.get("sterr") or props.get("surface_terrain"))
        total_price = self._to_float(props.get("valeurfonc") or props.get("valeur_fonciere"))
        sale_date = try_parse_date(props.get("datemut") or props.get("date_mutation"))

        if property_type != subject.property_type:
            return None
        valid_living_area = round(living_area, 2) if living_area is not None and living_area > 0 else None
        valid_total_price = round(total_price, 2) if total_price is not None and total_price > 0 else None
        price_per_sqm = None
        if valid_living_area is not None and valid_total_price is not None:
            price_per_sqm = round(valid_total_price / valid_living_area, 2)

        longitude, latitude = _geometry_center(raw.get("geometry", {}))
        distance_m = None
        if latitude is not None and longitude is not None:
            distance_m = haversine_distance_m(subject.latitude, subject.longitude, latitude, longitude)

        reverse = {}
        if longitude is not None and latitude is not None:
            try:
                reverse = reverse_geocode(longitude, latitude)
            except Exception:
                reverse = {}

        comparable_address = self._build_comparable_address(props, reverse)
        street_name = self._extract_street_name(props, reverse)

        return {
            "address": comparable_address or None,
            "property_type": property_type,
            "property_label": props.get("libtypbien") or props.get("type_local") or (property_type.title() if property_type else "Inconnu"),
            "living_area_sqm": valid_living_area,
            "rooms": self._extract_rooms(props, property_type),
            "land_area_sqm": round(land_area, 2) if land_area is not None else None,
            "sale_date": sale_date,
            "total_price_eur": valid_total_price,
            "price_per_sqm_eur": price_per_sqm,
            "street_name": street_name,
            "distance_m": round(distance_m, 1) if distance_m is not None else None,
            "similarity_score": None,
            "_micro_location_key": normalize_micro_location(street_name),
            "_latitude": latitude,
            "_longitude": longitude,
        }

    def _select_comparables(
        self,
        subject: SubjectProperty,
        comparables: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        return self._sort_comparables(comparables)[:MAX_RETAINED_COMPARABLES]

    @staticmethod
    def _deduplicate_recent_same_address_sales(comparables: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        kept_by_address: Dict[str, List[date]] = {}
        deduplicated: List[Dict[str, Any]] = []

        comparables_sorted = sorted(
            comparables,
            key=lambda comp: comp.get("sale_date") or date.min,
            reverse=True,
        )

        for comp in comparables_sorted:
            address = comp.get("address")
            sale_date = comp.get("sale_date")
            if not address or sale_date is None:
                deduplicated.append(comp)
                continue

            address_key = _normalize_address_text(str(address))
            kept_dates = kept_by_address.setdefault(address_key, [])
            if any(abs((kept_date - sale_date).days) < 365 for kept_date in kept_dates):
                continue

            kept_dates.append(sale_date)
            deduplicated.append(comp)

        return deduplicated

    @staticmethod
    def _sort_comparables(comparables: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        return sorted(
            comparables,
            key=lambda comp: (
                -(comp.get("similarity_score") or 0),
                -(comp["sale_date"].toordinal() if comp.get("sale_date") else 0),
                comp.get("distance_m") if comp.get("distance_m") is not None else float("inf"),
            ),
        )

    def _format_comparable_for_display(self, comp: Dict[str, Any], *, retained: bool) -> Dict[str, Any]:
        return {
            "Retenu": "Oui" if retained else "Non",
            "Score": format_french_number(comp.get("similarity_score")),
            "Adresse": comp.get("address"),
            "Type de bien": comp.get("property_label"),
            "Surface habitable": format_french_number(comp.get("living_area_sqm")),
            "Pièces": comp.get("rooms"),
            "Terrain extérieur": format_french_number(comp.get("land_area_sqm")),
            "Date de vente": comp["sale_date"].strftime("%d/%m/%Y") if comp.get("sale_date") else None,
            "Distance (m)": format_french_number(comp.get("distance_m")),
            "Prix de vente": format_french_number(comp.get("total_price_eur")),
            "Prix par m²": format_french_number(comp.get("price_per_sqm_eur")),
            "_latitude": comp.get("_latitude"),
            "_longitude": comp.get("_longitude"),
        }

    @staticmethod
    def _build_comparable_address(
        raw: Dict[str, Any],
        reverse: Dict[str, Any],
    ) -> Optional[str]:
        if reverse.get("label"):
            return str(reverse["label"]).strip()

        address_parts = [
            raw.get("l_num"),
            raw.get("l_typvoie"),
            raw.get("l_nomv"),
            raw.get("l_noma"),
            raw.get("nomvoie"),
            raw.get("adresse"),
        ]
        cleaned_parts = []
        for value in address_parts:
            if value is None:
                continue
            text = str(value).strip()
            if text and text not in cleaned_parts:
                cleaned_parts.append(text)

        if cleaned_parts:
            return " ".join(cleaned_parts)
        return None

    @staticmethod
    def _extract_street_name(
        raw: Dict[str, Any],
        reverse: Dict[str, Any],
    ) -> Optional[str]:
        for value in (
            reverse.get("street"),
            raw.get("l_nomv"),
            raw.get("l_noma"),
            raw.get("l_typvoie"),
            raw.get("nomvoie"),
            raw.get("lieudit"),
            raw.get("localite"),
        ):
            if value is None:
                continue
            text = str(value).strip()
            if text:
                return text
        return None

    @staticmethod
    def _normalize_property_type(raw: Dict[str, Any]) -> Optional[str]:
        value = str(raw.get("libtypbien") or raw.get("type_local") or "").strip().lower()
        if "appartement" in value:
            return "appartement"
        if "maison" in value:
            return "maison"
        apartment_count = ComparablePipeline._to_int(raw.get("nblocapt")) or 0
        house_count = ComparablePipeline._to_int(raw.get("nblocmai")) or 0
        apartment_area = ComparablePipeline._to_float(raw.get("sbatapt")) or 0
        house_area = ComparablePipeline._to_float(raw.get("sbatmai")) or 0
        if apartment_count > 0 and house_count == 0:
            return "appartement"
        if house_count > 0 and apartment_count == 0:
            return "maison"
        if apartment_area > 0 and house_area <= 0:
            return "appartement"
        if house_area > 0 and apartment_area <= 0:
            return "maison"
        return None

    @staticmethod
    def _to_float(value: Any) -> Optional[float]:
        if value is None or value == "":
            return None
        try:
            return float(str(value).replace(",", "."))
        except ValueError:
            return None

    @staticmethod
    def _to_int(value: Any) -> Optional[int]:
        if value is None or value == "":
            return None
        try:
            return int(float(str(value).replace(",", ".")))
        except ValueError:
            return None

    @classmethod
    def _extract_rooms(cls, raw: Dict[str, Any], property_type: Optional[str]) -> Optional[int]:
        for field_name in (
            "nbpprinc",
            "nombre_pieces_principales",
            "nombre_pieces",
            "nbpieces",
            "pieces",
        ):
            rooms = cls._to_int(raw.get(field_name))
            if rooms is not None and rooms > 0:
                return rooms

        if property_type not in {"appartement", "maison"}:
            return None

        prefix = "nbapt" if property_type == "appartement" else "nbmai"
        room_counts = {
            room_count: cls._to_int(raw.get(f"{prefix}{room_count}pp")) or 0
            for room_count in range(1, 6)
        }
        if sum(room_counts.values()) != 1:
            return None

        for room_count, unit_count in room_counts.items():
            if unit_count == 1:
                return room_count
        return None


def render_real_estate_tab():
    st.markdown("""
    <div class="step-card">
        <h3><span class="step-number">1</span> Bien à comparer</h3>
    </div>
    """, unsafe_allow_html=True)

    address_query = st.text_input(
        "Adresse",
        placeholder="Ex. 13 rue Victor Hugo",
        key="immo_address_query",
    )

    selected_address = address_query
    suggestions: List[Dict[str, Any]] = []

    if address_query and len(address_query.strip()) >= 3:
        try:
            suggestions = get_address_suggestions(address_query)
        except Exception as exc:
            st.warning(f"Impossible de charger les suggestions d'adresse : {exc}")

    if suggestions:
        options = [s["label"] for s in suggestions]
        selected_label = st.selectbox(
            "Suggestions d'adresses",
            options,
            index=None,
            placeholder="Choisis une suggestion si besoin",
            key="immo_address_selected",
        )
        if selected_label:
            selected_address = selected_label
    elif address_query:
        st.caption("Aucune suggestion trouvée. Tu peux quand même lancer la recherche avec l'adresse saisie.")

    property_type = st.selectbox(
        "Type de bien",
        ["appartement", "maison"],
        key="immo_property_type",
    )

    col1, col2 = st.columns(2)
    with col1:
        living_area_sqm = st.number_input(
            "Surface habitable (m²)",
            min_value=1.0,
            value=80.0,
            step=1.0,
            key="immo_surface",
        )
    with col2:
        rooms = st.number_input(
            "Nombre de pièces",
            min_value=1,
            max_value=20,
            value=4,
            step=1,
            key="immo_rooms",
        )

    land_area_sqm = None
    if property_type == "maison":
        land_area_sqm = st.number_input(
            "Surface terrain (m²)",
            min_value=0.0,
            value=100.0,
            step=10.0,
            key="immo_land_area",
        )

    st.markdown("""
    <div class="step-card">
        <h3><span class="step-number">2</span> Recherche</h3>
    </div>
    """, unsafe_allow_html=True)

    current_date = date.today()
    auto_api_min_year = subtract_months(current_date, DVF_MONTH_WINDOW).year
    current_api_min_year = st.session_state.get("immo_api_min_year", auto_api_min_year)
    st.session_state["immo_api_min_year"] = max(
        2000,
        min(int(current_api_min_year), auto_api_min_year),
    )

    col1, col2 = st.columns(2)
    with col1:
        search_radius_m = st.number_input(
            "Rayon de recherche (~m)",
            min_value=50,
            max_value=5000,
            value=DEFAULT_SEARCH_RADIUS_M,
            step=25,
            key="immo_search_radius_m",
        )
    with col2:
        api_min_year = st.number_input(
            "Année min API",
            min_value=2000,
            max_value=auto_api_min_year,
            value=st.session_state["immo_api_min_year"],
            step=1,
            key="immo_api_min_year",
        )

    launch = st.button(
        "Lancer comparatif",
        type="primary",
        use_container_width=True,
        key="launch_comparatif",
    )

    if not launch:
        st.info("Saisis une adresse, choisis le type de bien, puis lance le comparatif.")
        return

    if not selected_address or len(selected_address.strip()) < 5:
        st.error("Renseigne une adresse valide avant de lancer le comparatif.")
        return

    if property_type == "maison" and (land_area_sqm is None):
        st.error("Merci de renseigner la surface terrain pour une maison.")
        return

    try:
        payload = ComparableRequest(
            address=selected_address.strip(),
            property_type=property_type,
            living_area_sqm=living_area_sqm,
            rooms=rooms,
            land_area_sqm=land_area_sqm,
            search_radius_m=search_radius_m,
            api_min_year=api_min_year,
        )

        pipeline = ComparablePipeline(
            geocoder=GeocoderClient(),
            dvf_client=DVFClient(),
            scorer=ComparableScorer(),
        )

        with st.spinner("Comparatif immobilier en cours..."):
            result = pipeline.run(payload)
        st.session_state["immo_result"] = result

    except ValidationError as exc:
        st.error(f"Données invalides : {exc}")
        return
    except requests.exceptions.Timeout:
        st.error(
            "Le service immobilier a mis trop de temps à répondre. "
            "La requête a été relancée automatiquement, mais a fini en timeout. "
            "Réessaie dans quelques instants."
        )
        return
    except requests.exceptions.RequestException as exc:
        st.error(f"Erreur réseau vers les services immobiliers : {exc}")
        return
    except Exception as exc:
        st.error(f"Erreur : {exc}")
        return

    result = st.session_state.get("immo_result")
    if not result:
        return

    st.markdown("""
    <div class="step-card">
        <h3><span class="step-number">3</span> Bien cible</h3>
    </div>
    """, unsafe_allow_html=True)
    st.dataframe(
        build_subject_display_rows(result["subject"]),
        use_container_width=True,
        hide_index=True,
    )

    st.markdown("""
    <div class="step-card">
        <h3><span class="step-number">4</span> Comparables</h3>
    </div>
    """, unsafe_allow_html=True)

    comps = result.get("comparables", [])
    if not comps:
        st.warning("Aucun comparable trouvé. Vérifie le branchement DVF.")
        return

    display_comps = append_comparables_summary_row(comps, result.get("statistics", {}))
    st.dataframe(display_comps, use_container_width=True, hide_index=True)
    st.download_button(
        label="Télécharger le tableau Excel",
        data=build_immo_excel_export(result),
        file_name="comparatif_immobilier.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="download_immo_excel",
    )
