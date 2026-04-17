import argparse
import json
import logging
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_utils import (
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    HYPERLINK_FONT,
    LABEL_FONT,
    NUMBER_FORMAT_INTEGER,
    RADIATED_ROW_FILL,
    SECTION_FONT,
    THIN_BORDER,
    VALUE_FONT,
    apply_numeric_format,
    format_display_value,
    format_number_with_spaces,
    normalize_key,
    to_number,
)
from normalization import canonical_name, extract_person_folder, is_archived_path
from bilan_sheets import _build_bilan_sheet, _build_compte_resultat_sheet
from question_config import load_questions_config

logger = logging.getLogger(__name__)

_OPERATION_FIXED_ROWS_AFTER = {
    "montant_collecte": [
        ("Montant d'une obligation", "1"),
        ("Ticket Minimum", "1 000"),
    ]
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _normalize_folder_name(value: str) -> str:
    """Clé compacte (alphanumérique) basée sur la canonicalisation partagée."""
    return re.sub(r"[^a-z0-9]+", "", canonical_name(value or ""))




def _derive_person_folder_map_from_manifest(manifest_path: Path) -> Dict[str, str]:
    if not manifest_path.exists():
        return {}

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        return {}

    folder_map: Dict[str, str] = {}
    counter = 0
    for doc in manifest.get("files", []):
        folder_name = extract_person_folder(doc.get("source_path", ""))
        if not folder_name:
            continue
        key = _normalize_folder_name(folder_name)
        if key not in folder_map:
            folder_map[f"__{counter}"] = folder_name
            counter += 1

    return folder_map

def _merge_commentaires(*values) -> Optional[str]:
    merged: List[str] = []
    seen = set()

    for value in values:
        text = re.sub(r"\s+", " ", str(value or "").strip())
        if not text:
            continue

        dedupe_key = text.lower()
        if dedupe_key in seen:
            continue

        seen.add(dedupe_key)
        merged.append(text)

    return "\n".join(merged) or None
def _resolve_table_value(row_data: dict, key: str):
    alias_map = {
        "type_detention": ["type_detention", "type_de_detention", "type_bien"],
        "pct_detention": ["pct_detention"],
        "adresse": ["adresse", "adresse_destination"],
        "valeur_bien": ["valeur_bien", "estimation_actuelle"],
        "valeur_banque": ["valeur_banque", "capital_restant_du"],
        "valeur_nette_detenue": ["valeur_nette_detenue"],
        "revenus_locatifs": ["revenus_locatifs"],
        "garanties_donnees": ["garanties_donnees"],
        "societe": ["societe", "raison_sociale"],
        "activite": ["activite"],
        "creation": ["creation"],
        "infos_financieres": ["infos_financieres", "chiffres_cles"],
        "role_detention": ["role_detention", "pct_detention"],
        "commentaires": ["commentaires"],
    }

    candidates = alias_map.get(key, [key])
    normalized_row = {normalize_key(str(k)): v for k, v in row_data.items()}

    for candidate in candidates:
        if candidate in row_data and row_data[candidate] is not None:
            return row_data[candidate]
        normalized_candidate = normalize_key(candidate)
        if normalized_candidate in normalized_row and normalized_row[normalized_candidate] is not None:
            return normalized_row[normalized_candidate]

    if key == "valeur_nette_detenue":
        value_bien = to_number(_resolve_table_value(row_data, "valeur_bien"))
        value_banque = to_number(_resolve_table_value(row_data, "valeur_banque"))
        if value_bien is not None and value_banque is not None:
            net = value_bien - value_banque
            return int(net) if float(net).is_integer() else net

    return None


def _resolve_person_sheets(results: Dict, person_folder_map: Optional[Dict[str, str]] = None) -> List[Tuple[str, str]]:
    persons: List[Tuple[str, str]] = []

    if person_folder_map:
        for suffix, folder_name in sorted(person_folder_map.items()):
            if is_archived_path(folder_name):
                continue
            persons.append((suffix, folder_name))
        if persons:
            return persons

    for fid, value in sorted(results.items()):
        if fid.startswith("patrimoine_personne_nom") and value:
            suffix = fid.replace("patrimoine_personne_nom", "")
            persons.append((suffix, value.strip()))

    return persons


def _slugify_company_name(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "")
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value


def _build_pappers_company_url(company: dict) -> Optional[str]:
    nom_societe = (company.get("nom_societe") or company.get("societe") or "").strip()
    siren = str(company.get("siren") or "").strip()

    if not nom_societe or not siren:
        return None

    slug = _slugify_company_name(nom_societe)
    if not slug:
        return None

    return f"https://www.pappers.fr/entreprise/{slug}-{siren}"


def _build_infos_financieres(company: dict) -> Optional[str]:
    parts = []

    capital = company.get("capital")
    if capital is not None:
        parts.append(f"Capital: {format_number_with_spaces(capital)}")

    chiffre_affaires = company.get("chiffre_affaires")
    if chiffre_affaires is not None:
        parts.append(f"CA: {format_number_with_spaces(chiffre_affaires)}")

    resultat_net = company.get("resultat_net")
    if resultat_net is not None:
        parts.append(f"RN: {format_number_with_spaces(resultat_net)}")

    if not parts:
        return None

    return " | ".join(parts)


def _build_statut(company: dict) -> Optional[str]:
    parts = []

    statut = company.get("statut")
    if statut:
        parts.append(statut)

    statut_rcs = company.get("statut_rcs")
    if statut_rcs:
        parts.append(f"RCS: {statut_rcs}")

    if not parts:
        return None

    return " | ".join(parts)


def _is_radiated_company(company: dict) -> bool:
    statut = canonical_name(str(company.get("statut") or ""))
    statut_rcs = canonical_name(str(company.get("statut_rcs") or ""))
    return "radie" in statut or "radie" in statut_rcs


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def _build_operation_sheet(wb: Workbook, results: Dict, fields: List[Dict], ws=None):
    """Crée l'onglet Opération avec les champs simples, groupés par section."""
    ws = ws or wb.active
    ws.title = "Opération"

    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 50

    # Grouper les champs par section
    SECTIONS = [
        ("emprunt", "Société emprunteuse"),
        ("operation", "Société opération"),
    ]

    # Inclure les champs Opération + les champs sans sheet (localisation_projet, etc.)
    op_fields = [f for f in fields if f.get("excel_sheet", "") not in ("{person_name}", "{company_name}") and f.get("type") != "table"]

    row = 1
    filled = 0
    current_section = None

    for field in op_fields:
        field_id = field["field_id"]

        # Détecter le changement de section
        section_key = None
        for key, title in SECTIONS:
            if field_id.endswith(f"_{key}"):
                section_key = key
                break

        if section_key and section_key != current_section:
            if current_section is not None:
                row += 1  # ligne vide entre sections

            current_section = section_key
            section_title = dict(SECTIONS).get(section_key, "")
            section_cell = ws.cell(row=row, column=2, value=section_title)
            section_cell.font = Font(bold=True, size=12, color="FFFFFF")
            section_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            section_cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            row += 1

        # Champs hors section (objet_financement, taux, etc.)
        if section_key is None and current_section is not None:
            if current_section != "__other":
                row += 1
                other_cell = ws.cell(row=row, column=2, value="Opération")
                other_cell.font = Font(bold=True, size=12, color="FFFFFF")
                other_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                other_cell.alignment = Alignment(horizontal="center")
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
                row += 1
                current_section = "__other"

        label = field.get("label", field_id)
        value = results.get(field_id)

        ws.cell(row=row, column=2, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER

        val_cell = ws.cell(row=row, column=3, value=format_display_value(value))
        val_cell.font = VALUE_FONT
        val_cell.border = THIN_BORDER
        val_cell.alignment = Alignment(wrap_text=True)

        # Lien Google Earth pour la localisation
        if field_id == "localisation_projet" and value:
            earth_url = f"https://earth.google.com/web/search/{quote(value)}"
            val_cell.hyperlink = earth_url
            val_cell.font = HYPERLINK_FONT

        if value:
            filled += 1

        for fixed_label, fixed_value in _OPERATION_FIXED_ROWS_AFTER.get(field_id, []):
            row += 1

            ws.cell(row=row, column=2, value=fixed_label).font = LABEL_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER

            fixed_val_cell = ws.cell(row=row, column=3, value=fixed_value)
            fixed_val_cell.font = VALUE_FONT
            fixed_val_cell.border = THIN_BORDER
            fixed_val_cell.alignment = Alignment(wrap_text=True)

        row += 1

    return filled


def _build_patrimoine_sheet(wb: Workbook, results: Dict, fields: List[Dict],
                            person_folder_map: Optional[Dict[str, str]] = None):
    """Crée l'onglet Patrimoine avec toutes les personnes."""
    persons = _resolve_person_sheets(results, person_folder_map=person_folder_map)
    if not persons:
        return 0

    ws = wb.create_sheet("Patrimoine")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 25

    filled = 0
    current_row = 1

    # Récupérer les champs patrimoine
    simple_fields = [f for f in fields if f.get("excel_sheet") == "{person_name}" and f.get("type") != "table"]
    table_fields = [f for f in fields if f.get("excel_sheet") == "{person_name}" and f.get("type") == "table"]

    for person_idx, (suffix, person_name) in enumerate(persons):
        # Section header pour la personne
        section_cell = ws.cell(row=current_row, column=2, value=person_name)
        section_cell.font = Font(bold=True, size=13, color="2F5496")
        current_row += 1

        # Champs simples (ex: régime matrimonial)
        for field in simple_fields:
            field_id = field["field_id"]
            actual_fid = field_id + suffix
            actual_value = results.get(actual_fid)

            label = field.get("label", field_id)
            ws.cell(row=current_row, column=2, value=label).font = LABEL_FONT
            ws.cell(row=current_row, column=2).border = THIN_BORDER
            val_cell = ws.cell(row=current_row, column=3, value=format_display_value(actual_value))
            val_cell.font = VALUE_FONT
            val_cell.border = THIN_BORDER
            if actual_value:
                filled += 1
            current_row += 1

        # Tables (ex: patrimoine immobilier)
        for field in table_fields:
            field_id = field["field_id"]
            actual_fid = field_id + suffix
            actual_value = results.get(actual_fid)

            if not actual_value:
                continue

            try:
                rows = json.loads(actual_value) if isinstance(actual_value, str) else actual_value
            except (json.JSONDecodeError, TypeError):
                logger.warning(f"  {actual_fid}: valeur table non parseable")
                continue

            if not rows:
                continue

            current_row += 1
            table_label = field.get("label", field_id)
            ws.cell(row=current_row, column=2, value=table_label).font = SECTION_FONT
            current_row += 1

            col_map = field.get("column_mapping", {})
            # Insérer valeur_nette_detenue en colonne G si c'est le tableau patrimoine
            is_patrimoine_table = "valeur_bien" in col_map and "valeur_banque" in col_map
            display_headers = list(col_map.keys())
            if is_patrimoine_table and "valeur_nette_detenue" not in display_headers:
                # Insérer après valeur_banque (col F) → position G
                insert_idx = display_headers.index("valeur_banque") + 1 if "valeur_banque" in display_headers else len(display_headers)
                display_headers.insert(insert_idx, "valeur_nette_detenue")

            # Headers du tableau
            for col_idx, header in enumerate(display_headers):
                cell = ws.cell(row=current_row, column=2 + col_idx, value=header.replace("_", " ").title())
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
                cell.border = THIN_BORDER
            current_row += 1

            # Colonnes numériques pour la ligne TOTAL
            numeric_keys = {"valeur_bien", "valeur_banque", "valeur_nette_detenue", "revenus_locatifs"}

            # Données
            data_start_row = current_row
            for row_data in rows:
                for col_idx, key in enumerate(display_headers):
                    cell = ws.cell(row=current_row, column=2 + col_idx)
                    cell.border = THIN_BORDER

                    if key == "valeur_nette_detenue" and is_patrimoine_table:
                        col_e = col_map.get("valeur_bien", "E")
                        col_f = col_map.get("valeur_banque", "F")
                        col_c = col_map.get("pct_detention", "C")
                        r = current_row
                        cell.value = f"=({col_e}{r}-{col_f}{r})*{col_c}{r}"
                        cell.font = VALUE_FONT
                        cell.number_format = NUMBER_FORMAT_INTEGER
                    else:
                        cell_val = _resolve_table_value(row_data, key)
                        cell.value = format_display_value(cell_val) if not isinstance(cell_val, (int, float)) else cell_val
                        cell.font = VALUE_FONT
                        apply_numeric_format(cell, cell_val)
                        if cell_val is not None:
                            filled += 1
                current_row += 1
            data_end_row = current_row - 1

            # Ligne TOTAL
            if len(rows) > 0:
                total_font = Font(bold=True, size=10)
                ws.cell(row=current_row, column=2, value="TOTAL").font = total_font
                ws.cell(row=current_row, column=2).border = THIN_BORDER
                for col_idx, key in enumerate(display_headers):
                    cell = ws.cell(row=current_row, column=2 + col_idx)
                    cell.border = THIN_BORDER
                    if key in numeric_keys:
                        col_letter = get_column_letter(2 + col_idx)
                        cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                        cell.font = total_font
                        cell.number_format = NUMBER_FORMAT_INTEGER
                current_row += 1

        current_row += 2  # espace entre personnes

    return filled


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_mandats_sheet(wb: Workbook, pappers_mandats: Dict):
    """Crée l'onglet Mandats avec les sociétés Pappers."""
    if not pappers_mandats:
        return 0

    ws = wb.create_sheet("Mandats")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 30

    # Header
    headers = ["Société", "Rôle", "Activité", "Création", "Infos financières", "SIREN", "Statut", "Nb représentants", "Détention", "Commentaires"]
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=2 + col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    filled = 0
    current_row = 3

    for folder_name, folder_payload in pappers_mandats.items():
        if not folder_payload:
            continue

        # Titre du dossier
        folder_cell = ws.cell(row=current_row, column=2, value=folder_name)
        folder_cell.font = Font(bold=True, size=12, color="2F5496")
        current_row += 1

        def _write_company(row: int, company: dict) -> int:
            nonlocal filled

            nom = (company.get("nom_societe") or company.get("societe") or "").strip()
            statut_rcs = (company.get("statut_rcs") or "").strip().lower()
            row_fill = RADIATED_ROW_FILL if _is_radiated_company(company) else None
            role_value = company.get("role")
            if statut_rcs and statut_rcs != "inscrit" and not role_value:
                role_value = "X"

            cell = ws.cell(row=row, column=2, value=nom)
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

            url = _build_pappers_company_url(company)
            if url:
                cell.hyperlink = url
                cell.font = HYPERLINK_FONT
            else:
                cell.font = VALUE_FONT

            commentaires_value = _merge_commentaires(
                company.get("commentaires"),
                company.get("publication_difficulte_contenu"),
            )
            values = [
                role_value,
                company.get("activite"),
                company.get("date_creation") or company.get("creation"),
                _build_infos_financieres(company),
                str(company.get("siren") or "").strip() or None,
                _build_statut(company),
                company.get("nb_dirigeants_total"),
                company.get("detention"),
                commentaires_value,
            ]
            for col_idx, val in enumerate(values):
                display_val = format_display_value(val) if not isinstance(val, (int, float)) else val
                c = ws.cell(row=row, column=3 + col_idx, value=display_val)
                c.font = VALUE_FONT
                c.border = THIN_BORDER
                if col_idx == len(values) - 1:
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                if row_fill:
                    c.fill = row_fill
                apply_numeric_format(c, val)
                if val:
                    filled += 1

            if nom:
                filled += 1
            return row + 1

        if isinstance(folder_payload, list):
            for company in folder_payload:
                current_row = _write_company(current_row, company)
            total_cell = ws.cell(row=current_row, column=2, value=f"Total : {len(folder_payload)} société(s)")
            total_cell.font = Font(bold=True, size=10)
            current_row += 2
            continue

        if isinstance(folder_payload, dict):
            for person_name, companies in folder_payload.items():
                if not companies:
                    continue

                person_cell = ws.cell(row=current_row, column=2, value=person_name)
                person_cell.font = Font(bold=True, italic=True, size=10)
                current_row += 1

                for company in companies:
                    current_row = _write_company(current_row, company)

                total_cell = ws.cell(row=current_row, column=2, value=f"Total : {len(companies)} société(s)")
                total_cell.font = Font(bold=True, italic=True, size=9)
                current_row += 2

        current_row += 1

    return filled


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def fill_excel(results: Dict, fields: List[Dict], output_dir: Path,
               person_folder_map: Optional[Dict[str, str]] = None,
               pappers_mandats: Optional[Dict[str, List[dict]]] = None,
               include_operation: bool = True,
               include_patrimoine: bool = True,
               include_bilan: bool = True,
               include_compte_resultat: bool = True) -> Path:
    """Crée un Excel from scratch avec les données extraites."""
    output_path = output_dir / "rapport.xlsx"
    wb = Workbook()
    default_ws = wb.active

    filled = 0
    if include_operation:
        filled += _build_operation_sheet(wb, results, fields)
    else:
        wb.remove(default_ws)

    if include_patrimoine:
        filled += _build_patrimoine_sheet(wb, results, fields, person_folder_map=person_folder_map)
    if include_bilan:
        filled += _build_bilan_sheet(wb, results, fields)
    if include_compte_resultat:
        filled += _build_compte_resultat_sheet(wb, results)
    filled += _build_mandats_sheet(wb, pappers_mandats or {})

    if not wb.sheetnames:
        placeholder_ws = wb.create_sheet("Rapport")
        placeholder_ws["A1"] = "Aucun onglet selectionne."

    wb.save(str(output_path))
    logger.info("Excel genere: %s", output_path)
    return output_path


# Backward-compatible alias
def fill_excel_template(results: Dict, fields: List[Dict], template_path: str, output_dir: Path,
                        person_folder_map: Optional[Dict[str, str]] = None,
                        pappers_mandats: Optional[Dict[str, List[dict]]] = None):
    return fill_excel(results, fields, output_dir,
                      person_folder_map=person_folder_map,
                      pappers_mandats=pappers_mandats)


def main() -> None:
    parser = argparse.ArgumentParser(description="Crée un Excel depuis extraction_results.json")
    _root = Path(__file__).parent.parent.resolve()
    parser.add_argument(
        "--results",
        default=str(_root / "output" / "extraction_results.json"),
        help="Chemin vers extraction_results.json",
    )
    parser.add_argument(
        "--questions",
        default=None,
        help="Chemin vers un fichier de questions unique (sinon charge les fichiers split du dossier config)",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Dossier de sortie (par defaut: dossier du fichier extraction_results.json)",
    )
    args = parser.parse_args()

    results_path = Path(args.results)
    output_dir = Path(args.output_dir) if args.output_dir else results_path.parent
    manifest_path = output_dir / "manifest.json"

    if not results_path.exists():
        project_hint = output_dir.name if output_dir.name else None
        hint = (
            f" Lance d'abord `python run.py extract --project {project_hint}` puis relance `fill`."
            if project_hint
            else ""
        )
        raise FileNotFoundError(
            f"extraction_results.json introuvable: {results_path}.{hint}"
        )

    extraction_data = json.loads(results_path.read_text(encoding="utf-8"))
    questions_data = load_questions_config(_root / "config", explicit_path=args.questions)

    person_folder_map = extraction_data.get("person_folders")
    if not person_folder_map:
        person_folder_map = _derive_person_folder_map_from_manifest(manifest_path)

    mandats_results_path = output_dir / "mandats_results.json"

    pappers_mandats = None
    if mandats_results_path.exists():
        mandats_data = json.loads(mandats_results_path.read_text(encoding="utf-8"))
        pappers_mandats = mandats_data.get("societes_par_personne")

    fields = [
        f for f in questions_data["fields"]
        if isinstance(f, dict) and f.get("field_id")
    ]

    fill_excel(
        results=extraction_data["results"],
        fields=fields,
        output_dir=output_dir,
        person_folder_map=person_folder_map,
        pappers_mandats=pappers_mandats,
    )


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s — %(levelname)s — %(message)s",
    )
    main()
