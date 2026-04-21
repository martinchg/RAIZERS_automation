"""Construction de l'onglet Lots (grille de commercialisation)."""

import json
import logging
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from excel_utils import (
    BOLD_FONT,
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    NUMBER_FORMAT_INTEGER,
    THIN_BORDER,
    VALUE_FONT,
    apply_numeric_format,
    format_display_value,
    normalize_key,
)

logger = logging.getLogger(__name__)

_LOTS_NUMERIC_KEYS = {"surface_m2", "prix_m2", "prix_vente", "montant_depot_garantie"}

_LOTS_COLUMN_LABELS = {
    "lot_numero": "Lot n°",
    "niveau": "Niveau",
    "typologie": "Typologie",
    "surface_m2": "m²",
    "annexes": "Annexes",
    "prix_m2": "Prix/m²",
    "prix_vente": "Prix de vente",
    "precommercialise": "Précommercialisé",
    "etat_commercialisation": "Etat de commercialisation",
    "conditions_suspensives": "Conditions suspensives",
    "date_signature_contrat": "Date de signature",
    "montant_depot_garantie": "Dépôt de garantie",
}

_LOTS_TOTAL_KEYS = {"surface_m2", "prix_vente", "montant_depot_garantie"}


def _resolve_lot_value(row_data: dict, key: str):
    if key in row_data and row_data[key] is not None:
        return row_data[key]
    normalized_row = {normalize_key(str(k)): v for k, v in row_data.items()}
    normalized_key = normalize_key(key)
    if normalized_key in normalized_row:
        return normalized_row[normalized_key]
    return None


def build_lots_sheet(
    wb: Workbook,
    results: Dict,
    fields: List[Dict],
    logger_=None,
) -> int:
    lots_fields = [f for f in fields if f.get("field_id") == "lots_table"]
    if not lots_fields:
        return 0

    field = lots_fields[0]
    raw_value = results.get("lots_table")
    if not raw_value:
        return 0

    try:
        rows = json.loads(raw_value) if isinstance(raw_value, str) else raw_value
    except (json.JSONDecodeError, TypeError):
        (logger_ or logger).warning("lots_table: valeur non parseable")
        return 0

    if not rows:
        return 0

    col_map = field.get("column_mapping", {})
    display_keys = list(col_map.keys())

    ws = wb.create_sheet("Lots")

    col_widths = {
        "lot_numero": 10, "niveau": 10, "typologie": 12, "surface_m2": 8,
        "annexes": 20, "prix_m2": 12, "prix_vente": 14, "precommercialise": 16,
        "etat_commercialisation": 30, "conditions_suspensives": 25,
        "date_signature_contrat": 16, "montant_depot_garantie": 18,
    }
    for idx, key in enumerate(display_keys):
        ws.column_dimensions[get_column_letter(2 + idx)].width = col_widths.get(key, 15)

    current_row = 1
    title_cell = ws.cell(row=current_row, column=2, value="Grille de lots")
    title_cell.font = Font(bold=True, size=13, color="2F5496")
    current_row += 2

    for col_idx, key in enumerate(display_keys):
        cell = ws.cell(row=current_row, column=2 + col_idx, value=_LOTS_COLUMN_LABELS.get(key, key))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    current_row += 1

    filled = 0
    data_start_row = current_row

    for row_data in rows:
        for col_idx, key in enumerate(display_keys):
            cell = ws.cell(row=current_row, column=2 + col_idx)
            cell.border = THIN_BORDER
            value = _resolve_lot_value(row_data, key)
            if isinstance(value, (int, float)):
                cell.value = value
                cell.font = VALUE_FONT
                apply_numeric_format(cell, value)
            else:
                cell.value = format_display_value(value)
                cell.font = VALUE_FONT
            if value is not None:
                filled += 1
        current_row += 1

    data_end_row = current_row - 1

    ws.cell(row=current_row, column=2, value="TOTAL").font = BOLD_FONT
    ws.cell(row=current_row, column=2).border = THIN_BORDER
    for col_idx, key in enumerate(display_keys):
        cell = ws.cell(row=current_row, column=2 + col_idx)
        cell.border = THIN_BORDER
        if key in _LOTS_TOTAL_KEYS:
            col_letter = get_column_letter(2 + col_idx)
            cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
            cell.font = BOLD_FONT
            cell.number_format = NUMBER_FORMAT_INTEGER

    return filled
