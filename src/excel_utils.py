"""Utilitaires partages pour la generation Excel."""

import re
import unicodedata

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
RADIATED_ROW_FILL = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

LABEL_FONT = Font(bold=True, size=10)
VALUE_FONT = Font(size=10)
HYPERLINK_FONT = Font(size=10, color="0563C1", underline="single")
SECTION_FONT = Font(bold=True, size=11, color="2F5496")
RED_FONT = Font(bold=True, size=10, color="CC0000")
BOLD_FONT = Font(bold=True, size=10)
ITALIC_FONT = Font(italic=True, size=10)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

NUMBER_FORMAT_INTEGER = "# ##0"
NUMBER_FORMAT_DECIMAL = "# ##0.00"


def normalize_key(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(" ", "").replace("\u202f", "").replace("€", "").replace("%", "")
    text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def format_number_with_spaces(value) -> str:
    number = to_number(value)
    if number is None:
        return str(value)

    if isinstance(number, float) and not number.is_integer():
        text = f"{number:,.2f}".rstrip("0").rstrip(".")
    else:
        text = f"{int(number):,}"
    return text.replace(",", " ")


def format_display_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return format_number_with_spaces(value)

    text = str(value).strip()
    if not text:
        return ""
    if text.startswith("=") or "%" in text:
        return text
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", text):
        return text

    compact = text.replace(" ", "").replace("\u202f", "").replace("€", "")
    if re.fullmatch(r"-?\d+(?:[.,]\d+)?", compact):
        return format_number_with_spaces(text)

    return text


def apply_numeric_format(cell, value=None):
    target = cell.value if value is None else value
    number = to_number(target)
    if number is None:
        return
    cell.number_format = NUMBER_FORMAT_INTEGER if float(number).is_integer() else NUMBER_FORMAT_DECIMAL
