"""Tests pour l'extraction et la génération de l'onglet Lots."""

import json
import sys
from pathlib import Path

import pytest

SRC = Path(__file__).parent.parent / "src"
sys.path.insert(0, str(SRC))

from sheets.lots_sheet import build_lots_sheet
from openpyxl import Workbook


SAMPLE_LOTS = [
    {
        "lot_numero": "A01",
        "niveau": "RDC",
        "typologie": "T2",
        "surface_m2": 45.5,
        "annexes": "Balcon",
        "prix_m2": 4200,
        "prix_vente": 191100,
        "precommercialise": "non",
        "etat_commercialisation": "Disponible",
        "conditions_suspensives": None,
        "date_signature_contrat": None,
        "montant_depot_garantie": None,
    },
    {
        "lot_numero": "A02",
        "niveau": "R+1",
        "typologie": "T3",
        "surface_m2": 62.0,
        "annexes": "Terrasse",
        "prix_m2": 4300,
        "prix_vente": 266600,
        "precommercialise": "oui",
        "etat_commercialisation": "Signé - Dupont",
        "conditions_suspensives": "Prêt bancaire",
        "date_signature_contrat": "15/03/2025",
        "montant_depot_garantie": 10000,
    },
]

LOTS_FIELD = {
    "field_id": "lots_table",
    "label": "Grille de lots",
    "type": "table",
    "excel_sheet": "Lots",
    "column_mapping": {
        "lot_numero": "B",
        "niveau": "C",
        "typologie": "D",
        "surface_m2": "E",
        "annexes": "F",
        "prix_m2": "G",
        "prix_vente": "H",
        "precommercialise": "I",
        "etat_commercialisation": "J",
        "conditions_suspensives": "K",
        "date_signature_contrat": "L",
        "montant_depot_garantie": "M",
    },
}


def _make_wb_with_lots(raw_value):
    """Crée un workbook avec l'onglet Lots à partir de raw_value."""
    wb = Workbook()
    results = {"lots_table": raw_value}
    fields = [LOTS_FIELD]
    filled = build_lots_sheet(wb, results, fields)
    return wb, filled


class TestBuildLotsSheet:
    def test_onglet_cree_quand_json_string(self):
        """L'onglet Lots doit être créé quand lots_table est un JSON string."""
        raw = json.dumps(SAMPLE_LOTS, ensure_ascii=False)
        wb, filled = _make_wb_with_lots(raw)
        assert "Lots" in wb.sheetnames, "L'onglet Lots n'a pas été créé"
        assert filled > 0

    def test_onglet_cree_quand_liste(self):
        """L'onglet Lots doit être créé quand lots_table est déjà une liste."""
        wb, filled = _make_wb_with_lots(SAMPLE_LOTS)
        assert "Lots" in wb.sheetnames
        assert filled > 0

    def test_aucun_onglet_si_valeur_vide(self):
        """Aucun onglet ne doit être créé si lots_table est vide."""
        wb, filled = _make_wb_with_lots(None)
        assert "Lots" not in wb.sheetnames
        assert filled == 0

        wb2, filled2 = _make_wb_with_lots([])
        assert "Lots" not in wb2.sheetnames
        assert filled2 == 0

    def test_contenu_lignes(self):
        """Les données de chaque lot doivent être présentes dans les cellules."""
        raw = json.dumps(SAMPLE_LOTS, ensure_ascii=False)
        wb, _ = _make_wb_with_lots(raw)
        ws = wb["Lots"]

        # Ligne d'en-têtes = ligne 3, données à partir de la ligne 4
        values_col_B = [ws.cell(row=r, column=2).value for r in range(4, 7)]
        assert "A01" in values_col_B
        assert "A02" in values_col_B

    def test_ligne_total_presente(self):
        """La ligne TOTAL doit être présente sous les données."""
        raw = json.dumps(SAMPLE_LOTS, ensure_ascii=False)
        wb, _ = _make_wb_with_lots(raw)
        ws = wb["Lots"]

        total_row = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "TOTAL":
                    total_row = cell.row
                    break

        assert total_row is not None, "Ligne TOTAL introuvable"

    def test_stringify_non_table_ne_casse_pas_le_json(self):
        """
        Régression : avant le fix, stringify_non_table_value() était appelée sur la liste,
        produisant une chaîne du type 'A01, A02' non parseable par build_lots_sheet.
        On vérifie que json.loads() fonctionne sur la valeur stockée dans results.
        """
        from extraction.extract_structured_runtime import stringify_non_table_value

        raw_list = SAMPLE_LOTS
        stringified = stringify_non_table_value(raw_list)

        # La valeur stringifiée NE doit PAS être un JSON valide d'array
        try:
            parsed = json.loads(stringified)
            is_valid_lots_json = isinstance(parsed, list) and len(parsed) == len(SAMPLE_LOTS)
        except (json.JSONDecodeError, TypeError):
            is_valid_lots_json = False

        assert not is_valid_lots_json, (
            "stringify_non_table_value ne doit pas produire un JSON d'array — "
            "utiliser json.dumps pour les champs type=table"
        )

        # En revanche, json.dumps doit produire un JSON parseable
        dumped = json.dumps(raw_list, ensure_ascii=False)
        parsed2 = json.loads(dumped)
        assert isinstance(parsed2, list)
        assert len(parsed2) == 2
